r"""
Script 3: Options Flow Strength Analyser  (v3 — interactive thresholds + on-demand Layer 1)
═════════════════════════════════════════════════════════════════════════════════════════════
Reads oi_2026_06_16.csv and produces a multi-layer strength report.

v3 ADDITIONS (all v2 logic preserved):
  1. Layer 1 table is printed ONLY when user explicitly requests it
     (prompted at startup — "Show Layer 1 table? [y/N]")
  2. Layer 1 now includes MAX BEST MOVE column — the maximum favourable
     price move (in points) the option made AFTER the signal candle,
     within the same session.  Helps judge how good the signal quality
     was in hindsight.
  3. Interactive threshold prompts at startup:
       • STRENGTH_PCT  (default 60) — percentile for "strong"
       • WEAK_PCT      (default 40) — percentile for "weak"
       • Delta % min   (default 3 ) — minimum delta % change for entry
       • Gamma % min   (default 2 ) — minimum gamma % change for entry
     All four can be left blank to keep their defaults.

COMMAND LINE INPUT EXAMPLES
───────────────────────────
NOTE:
  Monthly/yearly modes also support the same entry filters:
  --delta-pct, --gama-pct/--gamma-pct, --volume-pct, --price-pct, --scr-above
  Debug more entries with:
  --top-n 100 --max-entries 20
  Price filter means the selected CE/PE strike option price must increase
  by at least --price-pct on the signal candle.
  Improved profit booking can be tuned with:
  --stop-loss, --trail-activate, --trail-lock, --trail-drop, --weak-exit-bars, --weak-exit-max-profit

Single CSV/date:
  python .\3_strength_analyser.py --csv oi_2026_06_09.csv --date 2026-06-02 --delta-pct 3 --gama-pct 2 --volume-pct 30 --price-pct 1 --scr-above 50

Debug more entries from option Greeks + option price move:
  python .\3_strength_analyser.py --csv oi_2024_10_29.csv --date 2024-10-29 --delta-pct 3 --gama-pct 2 --volume-pct 30 --price-pct 1 --scr-above 50 --top-n 100 --max-entries 20
  
Loose debug mode, show many entries and let weak signals fail:
  python .\3_strength_analyser.py --csv oi_2024_10_29.csv --date 2024-10-29 --delta-pct 1 --gama-pct 1 --volume-pct 10 --price-pct 0 --scr-above 0 --top-n 200 --max-entries 50

Monthly summary, Jan-Dec 2024:
  python .\3_strength_analyser.py --all-csv --csv-dir "..\..\niftyoptiondata" --monthly yes --year 2024 --from-month jan --to-month dec --delta-pct 3 --gama-pct 2 --volume-pct 30 --scr-above 50 --stop-loss 25 --trail-activate 15 --trail-lock 8 --trail-drop 12 --weak-exit-bars 4 --weak-exit-max-profit 5

Yearly summary, full 2024:
  python .\3_strength_analyser.py --all-csv --csv-dir "..\..\niftyoptiondata" --yearly yes --year 2024 --delta-pct 3 --gama-pct 2 --volume-pct 30 --scr-above 50

Convert DB files first, then monthly summary:
  python .\3_strength_analyser.py --convert-db --all-csv --db-dir "..\..\niftyoptiondata" --csv-dir "..\..\niftyoptiondata" --monthly yes --year 2024 --from-month jan --to-month dec --delta-pct 3 --gama-pct 2 --volume-pct 30 --scr-above 50

Show every available command-line input:
  python .\3_strength_analyser.py --help
"""

import argparse
import contextlib
import io
import re
import sqlite3
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import time as dtime

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

# ══════════════════════════════════════════════════════════════════════════════
# USER CONFIG / COMMAND-LINE DEFAULTS
# Command-line args override these values. Run:
#   python .\3_strength_analyser.py --help
#
# Most-used CLI overrides:
#   --csv oi_2024_10_29.csv
#   --date 2024-10-29
#   --delta-pct 3
#   --gama-pct 2
#   --volume-pct 30
#   --price-pct 1
#   --scr-above 50
#   --top-n 100
#   --max-entries 20
#   --monthly yes --year 2024 --from-month jan --to-month dec
#   --yearly yes --year 2024
# ══════════════════════════════════════════════════════════════════════════════
CSV_PATH         = "oi_2026_06_09.csv"  # CLI: --csv / --csv-path
ANALYSIS_DATE    = "2026-06-02"         # CLI: --date / --analysis-date
SIDE             = "both"          # "ce" | "pe" | "both"
CROSS_CANDLES    = 5
STRIKES_NEARBY   = 2
ATM_RANGE_POINTS = 100
STRIKE_STEP      = 50

# ── Thresholds — can be overridden interactively ─────────────────────────────
STRENGTH_PCT     = 60     # CLI: --strength-pct
WEAK_PCT         = 40     # CLI: --weak-pct
DELTA_MIN_PCT    = 3.0    # CLI: --delta-pct
GAMMA_MIN_PCT    = 2.0    # CLI: --gama-pct / --gamma-pct
VOLUME_MIN_PCT   = 30.0   # CLI: --volume-pct
PRICE_MIN_PCT    = 1.0    # CLI: --price-pct       (CE/PE strike option price % must increase)
SCR_ABOVE        = None   # CLI: --scr-above

SESSION_START    = dtime(9, 20)
SESSION_END      = dtime(15, 30)
OUT_CSV          = "strength_report.csv"
OUT_XLSX         = "strength_report.xlsx"
TOP_N            = 20      # CLI: --top-n 100       (print more candidate entries)

EXIT_NEG_COUNT   = 3
MIN_HOLD_BARS    = 12
EXIT_WEAK_BARS   = 4
TRAIL_DROP_PTS   = 12

ENTRY_CUTOFF     = dtime(15, 0)
FORCE_EXIT       = dtime(15, 30)
TRAIL_ACTIVATE_PTS = 15
TRAIL_LOCK_PTS     = 8
WEAK_EXIT_MAX_PROFIT_PTS = 5
STOP_LOSS_PTS      = 25
USE_STOP_PRICE_FILL = True

ENTRY_START                 = dtime(9, 20)
MAX_ENTRIES_PER_DAY         = 4  # CLI: --max-entries 20  (select more trades/day for debug)
MIN_GAP_BETWEEN_ENTRIES_MIN = 5

CHOP_WINDOW_CANDLES = 15
CHOP_RANGE_POINTS   = 20
OUT_CSV_SELECTED    = "selected_trades.csv"
DEFAULT_DATA_DIR    = Path(__file__).resolve().parents[2] / "niftyoptiondata"
SAVE_OUTPUTS        = True

# --- Run mode config (CLI args override these) ------------------------------
# Choose one:
#   "custom"              = use the detailed switches below
#   "single"              = run CSV_PATH + ANALYSIS_DATE only
#   "all_csv"             = run every oi_*.csv in CSV_DIR
#   "monthly"             = run selected year/month range and print monthly summary
#   "yearly"              = run selected full year and print yearly summary
#   "convert_and_all_csv" = convert DB files, then run every CSV
#   "convert_and_monthly" = convert DB files, then monthly summary
#   "convert_and_yearly"  = convert DB files, then yearly summary
#
# Examples:
#   RUN_MODE = "monthly"; BATCH_YEAR = 2024; FROM_MONTH = "jan"; TO_MONTH = "dec"
#   RUN_MODE = "yearly";  BATCH_YEAR = 2024
#   RUN_MODE = "convert_and_monthly"; BATCH_YEAR = 2024; FROM_MONTH = 1; TO_MONTH = 12
RUN_MODE = "custom"

PROMPT_USER_INPUT = False  # argparse/CLI mode by default; use --interactive for prompts
RUN_ALL_CSV      = False                  # CLI: --all-csv
CONVERT_DB       = False                  # CLI: --convert-db
CSV_DIR          = str(DEFAULT_DATA_DIR)  # CLI: --csv-dir
DB_DIR           = str(DEFAULT_DATA_DIR)  # CLI: --db-dir
OVERWRITE_CSV    = False                  # CLI: --overwrite-csv

MONTHLY_SUMMARY  = "n"     # CLI: --monthly yes
YEARLY_SUMMARY   = "n"     # CLI: --yearly yes
BATCH_YEAR       = None    # CLI: --year 2024
FROM_MONTH       = None    # CLI: --from-month jan
TO_MONTH         = None    # CLI: --to-month dec
FROM_DATE        = None    # CLI: --from-date 2024-01-01
TO_DATE          = None    # CLI: --to-date 2024-12-31

# ══════════════════════════════════════════════════════════════════════════════
# ANSI helpers
# ══════════════════════════════════════════════════════════════════════════════
def _c(t, code): return f"\033[{code}m{t}\033[0m"
def green(t):        return _c(t, "32")
def bright_green(t): return _c(t, "92")
def red(t):          return _c(t, "31")
def bright_red(t):   return _c(t, "91")
def yellow(t):       return _c(t, "33")
def orange(t):       return _c(t, "38;5;208")
def cyan(t):         return _c(t, "36")
def magenta(t):      return _c(t, "35")
def dim(t):          return _c(t, "2")
def bold(t):         return _c(t, "1")

def strip_ansi(s):
    return re.sub(r'\033\[[0-9;]*m', '', str(s))

def rjust(s, w):
    return ' ' * max(0, w - len(strip_ansi(s))) + s

def ljust(s, w):
    return s + ' ' * max(0, w - len(strip_ansi(s)))

def color_pct(val, extreme=20.0):
    if pd.isna(val): return dim("    n/a ")
    sign = "▲" if val > 0 else "▼" if val < 0 else " "
    txt  = f"{sign}{abs(val):6.2f}%"
    if val > extreme:    return bright_green(txt)
    elif val > 0:        return green(txt)
    elif val < -extreme: return bright_red(txt)
    elif val < 0:        return red(txt)
    return dim(txt)

def strength_bar(score, width=10):
    if score is None or pd.isna(score):
        return dim("[n/a]".ljust(width + 2))
    filled = int(round(float(score) / 100 * width))
    filled = max(0, min(width, filled))
    bar = "#" * filled + "-" * (width - filled)
    txt = f"[{bar}]"
    if score >= 75: return bright_green(txt)
    if score >= 50: return green(txt)
    if score >= 25: return yellow(txt)
    return red(txt)

def strength_label(score):
    if score >= 75: return bright_green("STRONG  ")
    if score >= 50: return green("MODERATE")
    if score >= 25: return yellow("WEAK    ")
    return red("VERY WK ")

def sep(w=130): print(dim("─" * w))
def header(title, w=130):
    print(bold("═" * w))
    print(bold(f"  {title}"))
    print(bold("═" * w))

def safe_float(x):
    try:
        v = float(x)
        return None if np.isnan(v) else v
    except Exception:
        return None

def apply_run_mode_defaults():
    global RUN_ALL_CSV, CONVERT_DB, MONTHLY_SUMMARY, YEARLY_SUMMARY

    mode = str(RUN_MODE or "custom").strip().lower()
    if mode == "custom":
        return

    RUN_ALL_CSV = False
    CONVERT_DB = False
    MONTHLY_SUMMARY = "n"
    YEARLY_SUMMARY = "n"

    if mode == "single":
        return
    if mode == "all_csv":
        RUN_ALL_CSV = True
    elif mode == "monthly":
        RUN_ALL_CSV = True
        MONTHLY_SUMMARY = "yes"
    elif mode == "yearly":
        RUN_ALL_CSV = True
        YEARLY_SUMMARY = "yes"
    elif mode == "convert_and_all_csv":
        CONVERT_DB = True
        RUN_ALL_CSV = True
    elif mode == "convert_and_monthly":
        CONVERT_DB = True
        RUN_ALL_CSV = True
        MONTHLY_SUMMARY = "yes"
    elif mode == "convert_and_yearly":
        CONVERT_DB = True
        RUN_ALL_CSV = True
        YEARLY_SUMMARY = "yes"
    else:
        raise ValueError(f"Unknown RUN_MODE: {RUN_MODE}")

def parse_cli_args():
    parser = argparse.ArgumentParser(
        description="Options flow strength analyser with CLI threshold overrides."
    )
    parser.add_argument("--csv", "--csv-path", dest="csv_path",
                        help="Input OI CSV file path.")
    parser.add_argument("--csv-dir", default=CSV_DIR,
                        help="Folder containing oi_*.csv files for batch mode.")
    parser.add_argument("--all-csv", action="store_true", default=RUN_ALL_CSV,
                        help="Run analyser for every oi_*.csv file in --csv-dir.")
    parser.add_argument("--convert-db", action="store_true", default=CONVERT_DB,
                        help="Convert oi_*.db files in --db-dir to CSV before running.")
    parser.add_argument("--db-dir", default=DB_DIR,
                        help="Folder containing oi_*.db files to convert.")
    parser.add_argument("--overwrite-csv", action="store_true", default=OVERWRITE_CSV,
                        help="Overwrite CSV files when converting DB files.")
    parser.add_argument("--monthly", choices=["y", "yes", "n", "no"], default=MONTHLY_SUMMARY,
                        help="Print month-by-month summary in batch mode.")
    parser.add_argument("--yearly", choices=["y", "yes", "n", "no"], default=YEARLY_SUMMARY,
                        help="Print year-by-year summary in batch mode.")
    parser.add_argument("--year", type=int, default=BATCH_YEAR,
                        help="Batch filter year, e.g. 2024.")
    parser.add_argument("--from-month", dest="from_month", default=FROM_MONTH,
                        help="Batch filter start month name/number, e.g. jan or 1.")
    parser.add_argument("--to-month", dest="to_month", default=TO_MONTH,
                        help="Batch filter end month name/number, e.g. dec or 12.")
    parser.add_argument("--from-date", dest="from_date", default=FROM_DATE,
                        help="Batch filter start date YYYY-MM-DD.")
    parser.add_argument("--to-date", dest="to_date", default=TO_DATE,
                        help="Batch filter end date YYYY-MM-DD.")
    parser.add_argument("--date", "--analysis-date", dest="analysis_date",
                        help="Analysis date in YYYY-MM-DD format.")
    parser.add_argument("--delta-pct", type=float, help="Minimum delta percent change for entry.")
    parser.add_argument("--gama-pct", "--gamma-pct", dest="gamma_pct", type=float,
                        help="Minimum gamma percent change for entry.")
    parser.add_argument("--volume-pct", type=float,
                        help="Minimum volume change percentage for entry.")
    parser.add_argument("--price-pct", type=float,
                        help="Minimum option price change percentage for entry.")
    parser.add_argument("--scr-above", type=float,
                        help="Only keep entries with computed entry score above this value.")
    parser.add_argument("--top-n", type=int, default=None,
                        help="How many top candidate entries to print in single-file mode.")
    parser.add_argument("--max-entries", type=int, default=None,
                        help="Maximum selected trades per day. Use a high value for debugging more entries.")
    parser.add_argument("--strength-pct", type=float,
                        help="Percentile threshold for same-candle strong spikes.")
    parser.add_argument("--weak-pct", type=float,
                        help="Percentile threshold for weak readings.")
    parser.add_argument("--stop-loss", type=float, default=None,
                        help="Hard stop-loss in option points from entry price.")
    parser.add_argument("--stop-price-fill", action="store_true", dest="stop_price_fill", default=None,
                        help="When stop/trail is touched, book exit at the stop level.")
    parser.add_argument("--market-fill", action="store_false", dest="stop_price_fill",
                        help="When stop/trail is touched, book exit at the current candle price.")
    parser.add_argument("--trail-activate", type=float, default=None,
                        help="Profit points before trailing stop activates.")
    parser.add_argument("--trail-lock", type=float, default=None,
                        help="Minimum points locked after trailing activates.")
    parser.add_argument("--trail-drop", type=float, default=None,
                        help="Exit if option falls this many points from best price after trailing activates.")
    parser.add_argument("--weak-exit-bars", type=int, default=None,
                        help="Consecutive weak bars needed for weak-before-profit exit.")
    parser.add_argument("--weak-exit-max-profit", type=float, default=None,
                        help="Allow weak-before-profit exit only if best move is <= this many points.")
    parser.add_argument("--show-layer1", action="store_true",
                        help="Print Layer 1 same-candle spike table.")
    parser.add_argument("--interactive", action="store_true", default=PROMPT_USER_INPUT,
                        help="Ask runtime input prompts for thresholds/layer1.")
    parser.add_argument("--no-interactive", action="store_false", dest="interactive",
                        help="Skip startup prompts and use config/CLI values.")
    return parser.parse_args()

def apply_cli_args(args):
    global CSV_PATH, ANALYSIS_DATE
    global CSV_DIR, DB_DIR, RUN_ALL_CSV, CONVERT_DB, OVERWRITE_CSV
    global MONTHLY_SUMMARY, YEARLY_SUMMARY, BATCH_YEAR, FROM_MONTH, TO_MONTH, FROM_DATE, TO_DATE
    global STRENGTH_PCT, WEAK_PCT, DELTA_MIN_PCT, GAMMA_MIN_PCT, VOLUME_MIN_PCT, PRICE_MIN_PCT, SCR_ABOVE
    global TRAIL_ACTIVATE_PTS, TRAIL_LOCK_PTS, TRAIL_DROP_PTS, EXIT_WEAK_BARS, WEAK_EXIT_MAX_PROFIT_PTS
    global STOP_LOSS_PTS, USE_STOP_PRICE_FILL
    global TOP_N, MAX_ENTRIES_PER_DAY

    if args.csv_path:
        CSV_PATH = args.csv_path
    if args.analysis_date:
        ANALYSIS_DATE = args.analysis_date
    CSV_DIR = args.csv_dir
    DB_DIR = args.db_dir
    RUN_ALL_CSV = args.all_csv
    CONVERT_DB = args.convert_db
    OVERWRITE_CSV = args.overwrite_csv
    MONTHLY_SUMMARY = args.monthly
    YEARLY_SUMMARY = args.yearly
    BATCH_YEAR = args.year
    FROM_MONTH = args.from_month
    TO_MONTH = args.to_month
    FROM_DATE = args.from_date
    TO_DATE = args.to_date
    if args.strength_pct is not None:
        STRENGTH_PCT = args.strength_pct
    if args.weak_pct is not None:
        WEAK_PCT = args.weak_pct
    if args.delta_pct is not None:
        DELTA_MIN_PCT = args.delta_pct
    if args.gamma_pct is not None:
        GAMMA_MIN_PCT = args.gamma_pct
    if args.volume_pct is not None:
        VOLUME_MIN_PCT = args.volume_pct
    if args.price_pct is not None:
        PRICE_MIN_PCT = args.price_pct
    if args.scr_above is not None:
        SCR_ABOVE = args.scr_above
    if args.top_n is not None:
        TOP_N = args.top_n
    if args.max_entries is not None:
        MAX_ENTRIES_PER_DAY = args.max_entries
    if args.stop_loss is not None:
        STOP_LOSS_PTS = args.stop_loss
    if args.stop_price_fill is not None:
        USE_STOP_PRICE_FILL = args.stop_price_fill
    if args.trail_activate is not None:
        TRAIL_ACTIVATE_PTS = args.trail_activate
    if args.trail_lock is not None:
        TRAIL_LOCK_PTS = args.trail_lock
    if args.trail_drop is not None:
        TRAIL_DROP_PTS = args.trail_drop
    if args.weak_exit_bars is not None:
        EXIT_WEAK_BARS = args.weak_exit_bars
    if args.weak_exit_max_profit is not None:
        WEAK_EXIT_MAX_PROFIT_PTS = args.weak_exit_max_profit

def has_cli_overrides(args):
    return any([
        args.csv_path is not None,
        args.all_csv,
        args.convert_db,
        args.analysis_date is not None,
        args.delta_pct is not None,
        args.gamma_pct is not None,
        args.volume_pct is not None,
        args.price_pct is not None,
        args.scr_above is not None,
        args.top_n is not None,
        args.max_entries is not None,
        args.stop_loss is not None,
        args.stop_price_fill is not None,
        args.trail_activate is not None,
        args.trail_lock is not None,
        args.trail_drop is not None,
        args.weak_exit_bars is not None,
        args.weak_exit_max_profit is not None,
        args.strength_pct is not None,
        args.weak_pct is not None,
        args.show_layer1,
        args.monthly in ("y", "yes"),
        args.yearly in ("y", "yes"),
        args.year is not None,
        args.from_month is not None,
        args.to_month is not None,
        args.from_date is not None,
        args.to_date is not None,
    ])

def print_active_options(show_layer1):
    print()
    print(dim("  --- Active thresholds ------------------------------------------------"))
    print(f"  {dim('CSV_PATH :')} {bold(str(CSV_PATH))}   "
          f"{dim('ANALYSIS_DATE :')} {bold(str(ANALYSIS_DATE))}")
    print(f"  {dim('CSV_DIR :')} {bold(str(CSV_DIR))}   "
          f"{dim('RUN_ALL_CSV :')} {bold(str(RUN_ALL_CSV))}   "
          f"{dim('CONVERT_DB :')} {bold(str(CONVERT_DB))}")
    print(f"  {dim('MONTHLY :')} {bold(str(MONTHLY_SUMMARY))}   "
          f"{dim('YEARLY :')} {bold(str(YEARLY_SUMMARY))}   "
          f"{dim('YEAR :')} {bold(str(BATCH_YEAR))}   "
          f"{dim('MONTHS :')} {bold(str(FROM_MONTH))} -> {bold(str(TO_MONTH))}")
    print(f"  {dim('STRENGTH_PCT :')} {bold(str(STRENGTH_PCT))}   "
          f"{dim('WEAK_PCT :')} {bold(str(WEAK_PCT))}   "
          f"{dim('DELTA_MIN% :')} {bold(str(DELTA_MIN_PCT))}   "
          f"{dim('GAMMA_MIN% :')} {bold(str(GAMMA_MIN_PCT))}   "
          f"{dim('VOLUME_MIN% :')} {bold(str(VOLUME_MIN_PCT))}   "
          f"{dim('PRICE_MIN% :')} {bold(str(PRICE_MIN_PCT))}   "
          f"{dim('SCR_ABOVE :')} {bold(str(SCR_ABOVE) if SCR_ABOVE is not None else 'OFF')}")
    print(f"  {dim('TOP_N :')} {bold(str(TOP_N))}   "
          f"{dim('MAX_ENTRIES_PER_DAY :')} {bold(str(MAX_ENTRIES_PER_DAY))}")
    print(f"  {dim('STOP_LOSS :')} {bold(str(STOP_LOSS_PTS))}   "
          f"{dim('STOP_FILL :')} {bold('STOP_PRICE' if USE_STOP_PRICE_FILL else 'MARKET_CANDLE')}   "
          f"{dim('TRAIL_ACTIVATE :')} {bold(str(TRAIL_ACTIVATE_PTS))}   "
          f"{dim('TRAIL_LOCK :')} {bold(str(TRAIL_LOCK_PTS))}   "
          f"{dim('TRAIL_DROP :')} {bold(str(TRAIL_DROP_PTS))}   "
          f"{dim('WEAK_EXIT_BARS :')} {bold(str(EXIT_WEAK_BARS))}   "
          f"{dim('WEAK_MAX_PROFIT :')} {bold(str(WEAK_EXIT_MAX_PROFIT_PTS))}")
    print(f"  {dim('Layer 1 table:')} {bold(bright_green('YES') if show_layer1 else dim('NO (skipped)'))}\n")

# ══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE STARTUP PROMPTS
# ══════════════════════════════════════════════════════════════════════════════
def ask_user_options():
    """
    Ask the user at startup:
      1. Whether to show the Layer 1 table
      2. Override thresholds (strength_pct, weak_pct, delta_min, gamma_min)

    Returns (show_layer1: bool, strength_pct, weak_pct, delta_min, gamma_min)
    """
    global STRENGTH_PCT, WEAK_PCT, DELTA_MIN_PCT, GAMMA_MIN_PCT, VOLUME_MIN_PCT, PRICE_MIN_PCT, SCR_ABOVE
    global TOP_N, MAX_ENTRIES_PER_DAY

    print(bold(cyan("\n  ╔══════════════════════════════════════════════════════╗")))
    print(bold(cyan("  ║     OPTIONS FLOW STRENGTH ANALYSER  v3              ║")))
    print(bold(cyan("  ╚══════════════════════════════════════════════════════╝\n")))

    # ── 1. Show Layer 1 table? ────────────────────────────────────────────────
    ans = input(bold("  Show Layer 1 (Same-Candle Spike) table? [y/N]: ")).strip().lower()
    show_layer1 = ans in ("y", "yes")

    print()
    print(dim("  ── Threshold overrides (press Enter to keep default) ──────────────"))

    # ── 2. STRENGTH_PCT ───────────────────────────────────────────────────────
    raw = input(bold(f"  Strength percentile threshold  [default {STRENGTH_PCT}]: ")).strip()
    if raw:
        try:
            v = float(raw)
            if 1 <= v <= 99:
                STRENGTH_PCT = v
                print(green(f"    → STRENGTH_PCT set to {STRENGTH_PCT}"))
            else:
                print(yellow(f"    ⚠  Out of range (1–99). Keeping default {STRENGTH_PCT}."))
        except ValueError:
            print(yellow(f"    ⚠  Invalid input. Keeping default {STRENGTH_PCT}."))

    # ── 3. WEAK_PCT ───────────────────────────────────────────────────────────
    raw = input(bold(f"  Weak   percentile threshold   [default {WEAK_PCT}]: ")).strip()
    if raw:
        try:
            v = float(raw)
            if 1 <= v <= 99:
                WEAK_PCT = v
                print(green(f"    → WEAK_PCT set to {WEAK_PCT}"))
            else:
                print(yellow(f"    ⚠  Out of range (1–99). Keeping default {WEAK_PCT}."))
        except ValueError:
            print(yellow(f"    ⚠  Invalid input. Keeping default {WEAK_PCT}."))

    # ── 4. DELTA_MIN_PCT ─────────────────────────────────────────────────────
    raw = input(bold(f"  Min Delta % change for entry   [default {DELTA_MIN_PCT}]: ")).strip()
    if raw:
        try:
            v = float(raw)
            DELTA_MIN_PCT = v
            print(green(f"    → DELTA_MIN_PCT set to {DELTA_MIN_PCT}"))
        except ValueError:
            print(yellow(f"    ⚠  Invalid input. Keeping default {DELTA_MIN_PCT}."))

    # ── 5. GAMMA_MIN_PCT ─────────────────────────────────────────────────────
    raw = input(bold(f"  Min Gamma % change for entry   [default {GAMMA_MIN_PCT}]: ")).strip()
    if raw:
        try:
            v = float(raw)
            GAMMA_MIN_PCT = v
            print(green(f"    → GAMMA_MIN_PCT set to {GAMMA_MIN_PCT}"))
        except ValueError:
            print(yellow(f"    ⚠  Invalid input. Keeping default {GAMMA_MIN_PCT}."))

    raw = input(bold(f"  Min Volume % change for entry  [default {VOLUME_MIN_PCT}]: ")).strip()
    if raw:
        try:
            VOLUME_MIN_PCT = float(raw)
            print(green(f"    -> VOLUME_MIN_PCT set to {VOLUME_MIN_PCT}"))
        except ValueError:
            print(yellow(f"    Invalid input. Keeping default {VOLUME_MIN_PCT}."))

    raw = input(bold(f"  Min Price % change for entry   [default {PRICE_MIN_PCT}]: ")).strip()
    if raw:
        try:
            PRICE_MIN_PCT = float(raw)
            print(green(f"    -> PRICE_MIN_PCT set to {PRICE_MIN_PCT}"))
        except ValueError:
            print(yellow(f"    Invalid input. Keeping default {PRICE_MIN_PCT}."))

    raw = input(bold("  Entry score must be above      [default OFF]: ")).strip()
    if raw:
        try:
            SCR_ABOVE = float(raw)
            print(green(f"    -> SCR_ABOVE set to {SCR_ABOVE}"))
        except ValueError:
            print(yellow("    Invalid input. Keeping SCR_ABOVE OFF."))

    raw = input(bold(f"  Top candidate rows to print    [default {TOP_N}]: ")).strip()
    if raw:
        try:
            TOP_N = max(1, int(raw))
            print(green(f"    -> TOP_N set to {TOP_N}"))
        except ValueError:
            print(yellow(f"    Invalid input. Keeping default {TOP_N}."))

    raw = input(bold(f"  Max selected trades per day    [default {MAX_ENTRIES_PER_DAY}]: ")).strip()
    if raw:
        try:
            MAX_ENTRIES_PER_DAY = max(1, int(raw))
            print(green(f"    -> MAX_ENTRIES_PER_DAY set to {MAX_ENTRIES_PER_DAY}"))
        except ValueError:
            print(yellow(f"    Invalid input. Keeping default {MAX_ENTRIES_PER_DAY}."))

    print_active_options(show_layer1)

    return show_layer1
def get_flow_hold_strength(full, ts, strike, side, col_map):
    """
    Returns flow strength for HOLD decision.
    Uses:
      1) same candle delta/gamma/volume/price
      2) cross-bar trend
      3) nearby strike confirmation
    """

    row_df = full[
        (full["timestamp"] == ts) &
        (full["strike"] == strike)
    ]

    if row_df.empty:
        return 0, "NO_ROW"

    r = row_df.iloc[0]

    same_points = sum([
        r[f"{side}_d_pct"] > 0,
        r[f"{side}_g_pct"] > 0,
        r[f"{side}_v_pct"] > 0,
        r[f"{side}_p_pct"] > 0,
    ])

    cross_points = sum([
        r[f"{side}_d_trend"] == 1,
        r[f"{side}_g_trend"] == 1,
        r[f"{side}_v_trend"] == 1,
        r[f"{side}_p_trend"] == 1,
    ])

    nearby = full[
        (full["timestamp"] == ts) &
        (full["strike"] >= strike - STRIKE_STEP * STRIKES_NEARBY) &
        (full["strike"] <= strike + STRIKE_STEP * STRIKES_NEARBY)
    ]

    nearby_strong = 0
    for _, nr in nearby.iterrows():
        ok = sum([
            nr[f"{side}_d_pct"] > 0,
            nr[f"{side}_g_pct"] > 0,
            nr[f"{side}_v_pct"] > 0,
            nr[f"{side}_p_pct"] > 0,
        ]) >= 3
        if ok:
            nearby_strong += 1

    cross_strike_ok = nearby_strong >= 2

    hold_score = same_points + cross_points + (2 if cross_strike_ok else 0)

    reason = (
        f"same={same_points}/4 "
        f"crossbar={cross_points}/4 "
        f"crossstrike={'YES' if cross_strike_ok else 'NO'}"
    )

    return hold_score, reason
# ══════════════════════════════════════════════════════════════════════════════
# EXIT FINDER
# ══════════════════════════════════════════════════════════════════════════════
def find_exit_after_entry(full, entry_ts, strike, side, entry_price, col_map):
    cm = col_map[side]

    trade = full[
        (full["timestamp"] > entry_ts) &
        (full["strike"] == strike)
    ].copy().sort_values("timestamp").reset_index(drop=True)

    if trade.empty:
        return entry_ts, entry_price, 0, "No future data"

    best_price = entry_price
    weak_count = 0
    flow_weak_count = 0
    hard_sl = max(0.0, entry_price - STOP_LOSS_PTS)

    for i, r in trade.iterrows():
        price = r[cm["price"]]

        if r["timestamp"].time() >= FORCE_EXIT:
            pnl = price - entry_price
            return r["timestamp"], price, pnl, "Exit: 3:30 PM force exit"

        best_price = max(best_price, price)

        if price <= hard_sl:
            exit_price = hard_sl if USE_STOP_PRICE_FILL else price
            pnl = exit_price - entry_price
            return r["timestamp"], exit_price, pnl, "Exit: hard stop loss"

        if i < MIN_HOLD_BARS:
            continue

        d_weak = r[f"{side}_d_pct"] < 0
        g_weak = r[f"{side}_g_pct"] < 0
        v_weak = r[f"{side}_v_pct"] < 0
        p_weak = r[f"{side}_p_pct"] < 0

        weak_now = sum([d_weak, g_weak, v_weak, p_weak]) >= 3

        if weak_now:
            weak_count += 1
        else:
            weak_count = 0

        if best_price >= entry_price + TRAIL_ACTIVATE_PTS:
            hold_score, hold_reason = get_flow_hold_strength(
                full, r["timestamp"], strike, side, col_map
            )

            flow_strong = hold_score >= 7
            flow_weak = hold_score <= 4

            if flow_strong:
                trail_sl = max(entry_price + TRAIL_LOCK_PTS, best_price - (TRAIL_DROP_PTS * 2))
            else:
                trail_sl = max(entry_price + TRAIL_LOCK_PTS, best_price - TRAIL_DROP_PTS)

            if flow_weak:
                flow_weak_count += 1
            else:
                flow_weak_count = 0

            if price <= trail_sl and flow_weak_count >= 3:
                exit_price = trail_sl if USE_STOP_PRICE_FILL else price
                pnl = exit_price - entry_price
                return r["timestamp"], exit_price, pnl, f"Exit: 3-bar flow weak + trail ({hold_reason})"

        best_profit = best_price - entry_price

        if weak_count >= EXIT_WEAK_BARS and best_profit <= WEAK_EXIT_MAX_PROFIT_PTS:
            pnl = price - entry_price
            return r["timestamp"], price, pnl, "Exit: weak before profit"

    last = trade.iloc[-1]
    exit_price = last[cm["price"]]
    pnl = exit_price - entry_price
    return last["timestamp"], exit_price, pnl, "Exit: session end"


# ══════════════════════════════════════════════════════════════════════════════
# MAX BEST MOVE  — how many points did this option gain AFTER the signal candle
# ══════════════════════════════════════════════════════════════════════════════
def compute_best_move_info(full, entry_ts, strike, side, entry_price, col_map):
    """
    Returns the maximum price the option reached AFTER the signal candle
    minus the entry price — i.e. the best possible exit in hindsight.

    This is informational only (not used for trade logic).
    Returns move/time/price info so missed moves can be audited.
    """
    cm = col_map[side]
    future = full[
        (full["timestamp"] > entry_ts) &
        (full["strike"] == strike)
    ][["timestamp", cm["price"]]].copy()

    if future.empty:
        return {
            "best_move_pnl": 0.0,
            "best_move_time": None,
            "best_move_price": entry_price,
        }

    best_idx = future[cm["price"]].idxmax()
    best_row = future.loc[best_idx]
    best_price = float(best_row[cm["price"]])
    return {
        "best_move_pnl": round(best_price - entry_price, 2),
        "best_move_time": best_row["timestamp"],
        "best_move_price": round(best_price, 2),
    }


def compute_max_best_move(full, entry_ts, strike, side, entry_price, col_map):
    return compute_best_move_info(full, entry_ts, strike, side, entry_price, col_map)["best_move_pnl"]


# ══════════════════════════════════════════════════════════════════════════════
# CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def pct_change_col(series):
    return (series.pct_change(fill_method=None) * 100).round(2)

def rolling_pct_rank(series, window=20):
    return series.rolling(window, min_periods=1).apply(
        lambda x: (x[-1] > x[:-1]).mean() * 100 if len(x) > 1 else 50.0,
        raw=True
    )

def compute_strength_score(delta_rank, gamma_rank, volume_rank, price_rank):
    return (
        0.30 * delta_rank +
        0.30 * gamma_rank +
        0.25 * volume_rank +
        0.15 * price_rank
    ).round(1)

def cross_candle_trend(series, n):
    def slope_sign(x):
        if len(x) < 2: return 0
        xs = np.arange(len(x))
        slope = np.polyfit(xs, x, 1)[0]
        if slope > 0:  return 1
        if slope < 0:  return -1
        return 0
    return series.rolling(n, min_periods=2).apply(slope_sign, raw=True)

def cross_candle_pct(series, n):
    return ((series - series.shift(n)) / series.shift(n).abs() * 100).round(2)

def is_choppy_market(ts, spot_series, window=CHOP_WINDOW_CANDLES, threshold=CHOP_RANGE_POINTS):
    hist = spot_series.loc[:ts].tail(window)
    if len(hist) < max(5, window // 2):
        return False
    return (hist.max() - hist.min()) < threshold


# ══════════════════════════════════════════════════════════════════════════════
# XLSX EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def _export_xlsx(full, layer1_rows, sides, col_map, save_cols, timestamps,
                 selected_trades=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XLSide
    from openpyxl.utils import get_column_letter

    HDR_BG  = PatternFill("solid", start_color="1F4E79")
    HDR_FONT = Font(name="Consolas", bold=True, color="FFFFFF", size=10)

    C_BRIGHT_GREEN = "92FF92"
    C_GREEN        = "00CC00"
    C_BRIGHT_RED   = "FF9292"
    C_RED          = "FF4444"
    C_YELLOW       = "FFFF00"
    C_CYAN         = "00FFFF"
    C_DIM          = "888888"
    C_WHITE        = "FFFFFF"

    thin = XLSide(style="thin", color="333355")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def pct_color(val, extreme=20.0):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return "n/a", C_DIM
        sign = "▲" if val > 0 else "▼" if val < 0 else " "
        txt  = f"{sign}{abs(val):.2f}%"
        if val > extreme:    clr = C_BRIGHT_GREEN
        elif val > 0:        clr = C_GREEN
        elif val < -extreme: clr = C_BRIGHT_RED
        elif val < 0:        clr = C_RED
        else:                clr = C_DIM
        return txt, clr

    def score_color(score):
        if score >= 75: return C_BRIGHT_GREEN
        if score >= 50: return C_GREEN
        if score >= 25: return C_YELLOW
        return C_RED

    def score_label(score):
        if score >= 75: return "STRONG"
        if score >= 50: return "MODERATE"
        if score >= 25: return "WEAK"
        return "VERY WK"

    def write_cell(ws, row, col, value, fg=C_WHITE, bg=None, bold=False, align="center"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(name="Consolas", color=fg, bold=bold, size=9)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = bdr
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        return cell

    def style_header_row(ws, cols, row=1):
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.font      = HDR_FONT
            cell.fill      = HDR_BG
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
        ws.row_dimensions[row].height = 18

    def autofit(ws, extra=3, max_w=35):
        for col_cells in ws.columns:
            best = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(best + extra, max_w)

    wb = Workbook()

    # ── Sheet 1: Layer 1 (now includes MAX BEST MOVE) ────────────────────────
    ws1 = wb.active
    ws1.title = "L1 Same-Candle Spikes"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "FFEB9C"

    l1_hdr = ["TIMESTAMP","STRIKE","SPOT","SIDE","SCORE","STRENGTH",
              "DELTA","DELTA_%","GAMMA","GAMMA_%",
              "VOLUME","VOLUME_%","PRICE","PRICE_%","IV",
              "MAX BEST MOVE"]   # ← NEW COLUMN
    style_header_row(ws1, l1_hdr)

    # layer1_rows now carries max_best_move as last element
    for ri, row in enumerate(layer1_rows, 2):
        ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv, max_move = row
        row_bg = "1A3A1A" if side == "ce" else "3A1A1A"
        sc     = score_color(score)
        sl     = score_label(score)
        dp_t, dp_c = pct_color(safe_float(dp))
        gp_t, gp_c = pct_color(safe_float(gp))
        vp_t, vp_c = pct_color(safe_float(vp), 50)
        pp_t, pp_c = pct_color(safe_float(pp))

        move_clr = C_BRIGHT_GREEN if max_move > 0 else C_DIM

        write_cell(ws1, ri, 1,  str(ts)[:19],  C_DIM,  row_bg)
        write_cell(ws1, ri, 2,  round(stk,1),  C_CYAN, row_bg)
        write_cell(ws1, ri, 3,  round(spot,2), C_CYAN, row_bg)
        write_cell(ws1, ri, 4,  side.upper(),  C_BRIGHT_GREEN if side=="ce" else C_BRIGHT_RED, row_bg, bold=True)
        write_cell(ws1, ri, 5,  round(score,1),sc,     row_bg, bold=True)
        write_cell(ws1, ri, 6,  sl,            sc,     row_bg)
        write_cell(ws1, ri, 7,  round(float(dv),4) if safe_float(dv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 8,  dp_t, dp_c, row_bg)
        write_cell(ws1, ri, 9,  round(float(gv),6) if safe_float(gv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 10, gp_t, gp_c, row_bg)
        write_cell(ws1, ri, 11, int(vv) if safe_float(vv) is not None else None, C_DIM, row_bg)
        write_cell(ws1, ri, 12, vp_t, vp_c, row_bg)
        write_cell(ws1, ri, 13, round(float(pv),2) if safe_float(pv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 14, pp_t, pp_c, row_bg)
        write_cell(ws1, ri, 15, round(float(iv),2) if safe_float(iv) is not None else None, C_DIM, row_bg)
        write_cell(ws1, ri, 16, f"+{max_move:.2f}" if max_move > 0 else f"{max_move:.2f}", move_clr, row_bg, bold=True)

    ws1.freeze_panes = "A2"
    autofit(ws1)

    # ── Sheet 2: Full Data ────────────────────────────────────────────────────
    ws_fd = wb.create_sheet("Full Data")
    ws_fd.sheet_view.showGridLines = False
    ws_fd.sheet_properties.tabColor = "444444"

    existing = [c for c in save_cols if c in full.columns]
    style_header_row(ws_fd, existing)

    for ri2, row in enumerate(full[existing].itertuples(index=False), 2):
        bg = "1A1A2E" if ri2 % 2 == 0 else "16213E"
        for ci2, val in enumerate(row, 1):
            if hasattr(val, "isoformat"):
                v = val.isoformat()
            elif isinstance(val, (bool, np.bool_)):
                v = str(val)
            elif isinstance(val, float) and np.isnan(val):
                v = None
            else:
                v = val
            write_cell(ws_fd, ri2, ci2, v, C_WHITE, bg)

    ws_fd.freeze_panes = "A2"
    autofit(ws_fd)

    # ── Sheet 3: Selected Trades ──────────────────────────────────────────────
    if selected_trades:
        ws6 = wb.create_sheet("Selected Trades (Final)")
        ws6.sheet_view.showGridLines = False
        ws6.sheet_properties.tabColor = "00FF7F"

        l6_hdr = ["ENTRY TIME","EXIT TIME","STRIKE","SPOT","SIDE",
                  "ENTRY","EXIT","PNL","BEST PNL","ENTRY SCORE","SYS SCORE",
                  "D%","G%","V%","P%","EXIT REASON"]
        style_header_row(ws6, l6_hdr)

        for ri, (ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points,
                 best_move_pnl, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason) in enumerate(selected_trades, 2):
            row_bg = "1A3A1A" if pnl_points >= 0 else "3A1A1A"
            dp_t, dp_c = pct_color(safe_float(dp))
            gp_t, gp_c = pct_color(safe_float(gp))
            vp_t, vp_c = pct_color(safe_float(vp), 50)
            pp_t, pp_c = pct_color(safe_float(pp))

            write_cell(ws6, ri, 1,  str(ts)[:19],         C_DIM, row_bg)
            write_cell(ws6, ri, 2,  str(exit_ts)[:19],    C_DIM, row_bg)
            write_cell(ws6, ri, 3,  round(stk,1),         C_CYAN, row_bg)
            write_cell(ws6, ri, 4,  round(spot,2),        C_CYAN, row_bg)
            write_cell(ws6, ri, 5,  side.upper(),         C_BRIGHT_GREEN if side=="ce" else C_BRIGHT_RED, row_bg, bold=True)
            write_cell(ws6, ri, 6,  round(entry_price,2), C_WHITE, row_bg)
            write_cell(ws6, ri, 7,  round(exit_price,2),  C_WHITE, row_bg)
            write_cell(ws6, ri, 8,  round(pnl_points,2),  C_BRIGHT_GREEN if pnl_points>=0 else C_BRIGHT_RED, row_bg, bold=True)
            write_cell(ws6, ri, 9,  round(best_move_pnl,2), C_BRIGHT_GREEN if best_move_pnl>=0 else C_BRIGHT_RED, row_bg, bold=True)
            write_cell(ws6, ri, 10, round(entry_score,1), C_YELLOW, row_bg, bold=True)
            write_cell(ws6, ri, 11, round(sys_score,1),   score_color(sys_score), row_bg)
            write_cell(ws6, ri, 12, dp_t, dp_c, row_bg)
            write_cell(ws6, ri, 13, gp_t, gp_c, row_bg)
            write_cell(ws6, ri, 14, vp_t, vp_c, row_bg)
            write_cell(ws6, ri, 15, pp_t, pp_c, row_bg)
            write_cell(ws6, ri, 16, exit_reason, C_WHITE, row_bg)

        ws6.freeze_panes = "A2"
        autofit(ws6)

    wb.save(OUT_XLSX)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def resolve_csv_path(csv_path):
    path = Path(csv_path)
    candidates = [path]
    if not path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        candidates += [
            script_dir / path,
            DEFAULT_DATA_DIR / path.name,
            Path(CSV_DIR) / path.name,
        ]

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return path


def run_single_analysis(show_layer1):
    global CSV_PATH

    csv = resolve_csv_path(CSV_PATH)
    if not csv.exists():
        raise FileNotFoundError(
            f"CSV not found: {CSV_PATH}\n"
            f"Tried current folder, script folder, CSV_DIR, and {DEFAULT_DATA_DIR}.\n"
            "Run 1_db_to_csv.py first if the CSV has not been converted yet."
        )
    CSV_PATH = str(csv)

    print(bold(cyan(f"  Loading {CSV_PATH} …")))
    df = pd.read_csv(csv, parse_dates=["timestamp"])
    print("CSV date range:", df["timestamp"].min(), "to", df["timestamp"].max())
    print("Available dates:", sorted(df["timestamp"].dt.date.unique())[:20])

    df = df[df["timestamp"].dt.date == pd.to_datetime(ANALYSIS_DATE).date()].copy()
    df.reset_index(drop=True, inplace=True)

    if df.empty:
        print(red(f"No rows found for ANALYSIS_DATE = {ANALYSIS_DATE}"))
        print(yellow("Use one date from Available dates printed above."))
        return

    df.sort_values(["timestamp", "strike"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    before = len(df)
    df = df[df["timestamp"].dt.time.between(SESSION_START, SESSION_END)].copy()
    df.reset_index(drop=True, inplace=True)

    df["atm_strike"] = (df["spot"] / STRIKE_STEP).round() * STRIKE_STEP
    df = df[
        (df["strike"] >= df["atm_strike"] - ATM_RANGE_POINTS) &
        (df["strike"] <= df["atm_strike"] + ATM_RANGE_POINTS)
    ].copy()
    df.drop(columns=["atm_strike"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    after = len(df)

    print(f"  {dim('Time filter:')} {SESSION_START.strftime('%H:%M')} – {SESSION_END.strftime('%H:%M')}  "
          f"{dim('Rows:')} {before:,} → {bright_green(str(after))}")

    sides   = ["ce","pe"] if SIDE == "both" else [SIDE]
    col_map = {
        "ce": {"delta":"call_delta","gamma":"call_gamma",
               "volume":"call_volume","price":"call_price","iv":"call_iv"},
        "pe": {"delta":"put_delta", "gamma":"put_gamma",
               "volume":"put_volume","price":"put_price","iv":"put_iv"},
    }

    print(f"  {dim('Timestamps:')} {df['timestamp'].nunique()}  "
          f"{dim('Strikes:')} {df['strike'].nunique()}  "
          f"{dim('Sides:')} {SIDE.upper()}  "
          f"{dim('X-candle window:')} {CROSS_CANDLES}  "
          f"{dim('Nearby strikes:')} ±{STRIKES_NEARBY}\n")

    # ── Feature engineering ───────────────────────────────────────────────────
    all_strikes = sorted(df["strike"].unique())
    records     = []

    for strike in all_strikes:
        sdf = df[df["strike"] == strike].copy().sort_values("timestamp").reset_index(drop=True)
        if len(sdf) < 3:
            continue

        for side in sides:
            cm = col_map[side]
            d  = sdf[cm["delta"]].copy()
            g  = sdf[cm["gamma"]].copy()
            v  = sdf[cm["volume"]].copy()
            p  = sdf[cm["price"]].copy()

            sdf[f"{side}_d_pct"]   = pct_change_col(d)
            sdf[f"{side}_g_pct"]   = pct_change_col(g)
            sdf[f"{side}_v_pct"]   = pct_change_col(v)
            sdf[f"{side}_p_pct"]   = pct_change_col(p)

            sdf[f"{side}_d_xpct"]  = cross_candle_pct(d, CROSS_CANDLES)
            sdf[f"{side}_g_xpct"]  = cross_candle_pct(g, CROSS_CANDLES)
            sdf[f"{side}_v_xpct"]  = cross_candle_pct(v, CROSS_CANDLES)
            sdf[f"{side}_p_xpct"]  = cross_candle_pct(p, CROSS_CANDLES)

            sdf[f"{side}_d_trend"] = cross_candle_trend(d, CROSS_CANDLES)
            sdf[f"{side}_g_trend"] = cross_candle_trend(g, CROSS_CANDLES)
            sdf[f"{side}_v_trend"] = cross_candle_trend(v, CROSS_CANDLES)
            sdf[f"{side}_p_trend"] = cross_candle_trend(p, CROSS_CANDLES)

            sdf[f"{side}_d_rank"]  = rolling_pct_rank(d)
            sdf[f"{side}_g_rank"]  = rolling_pct_rank(g)
            sdf[f"{side}_v_rank"]  = rolling_pct_rank(v)
            sdf[f"{side}_p_rank"]  = rolling_pct_rank(p)

            sdf[f"{side}_score"]   = compute_strength_score(
                sdf[f"{side}_d_rank"], sdf[f"{side}_g_rank"],
                sdf[f"{side}_v_rank"], sdf[f"{side}_p_rank"]
            )

            # Use RUNTIME STRENGTH_PCT (possibly overridden by user)
            sdf[f"{side}_same_spike"] = (
                (sdf[f"{side}_d_rank"] >= STRENGTH_PCT) &
                (sdf[f"{side}_g_rank"] >= STRENGTH_PCT)
            )
            sdf[f"{side}_cross_bull"] = (
                (sdf[f"{side}_d_trend"] == 1) &
                (sdf[f"{side}_g_trend"] == 1) &
                (sdf[f"{side}_v_trend"] == 1)
            )
            sdf[f"{side}_cross_bear"] = (
                (sdf[f"{side}_d_trend"] == -1) &
                (sdf[f"{side}_g_trend"] == -1) &
                (sdf[f"{side}_v_trend"] == -1)
            )

        records.append(sdf)

    full       = pd.concat(records, ignore_index=True).sort_values(["timestamp","strike"])
    timestamps = sorted(full["timestamp"].unique())

    spot_by_ts = (
        full.drop_duplicates(subset="timestamp")
            .set_index("timestamp")["spot"]
            .sort_index()
    )

    W = 158   # wider to accommodate MAX BEST MOVE column

    # ══════════════════════════════════════════════════════════════════════════
    # Build Layer 1 rows (always computed, but table printed only if requested)
    # ══════════════════════════════════════════════════════════════════════════
    layer1_rows = []
    for side in sides:
        sub = full[full[f"{side}_same_spike"]].copy()
        for _, r in sub.iterrows():
            # Compute max best move for this signal
            max_move = compute_max_best_move(
                full, r["timestamp"], r["strike"], side,
                r[col_map[side]["price"]], col_map
            )
            layer1_rows.append((
                r["timestamp"], r["strike"], r["spot"], side,
                r[f"{side}_score"],
                r[col_map[side]["delta"]], r[f"{side}_d_pct"],
                r[col_map[side]["gamma"]], r[f"{side}_g_pct"],
                r[col_map[side]["volume"]], r[f"{side}_v_pct"],
                r[col_map[side]["price"]], r[f"{side}_p_pct"],
                r[col_map[side]["iv"]],
                max_move   # ← NEW last element
            ))

    layer1_rows.sort(key=lambda x: (x[0], 0 if x[3] == "ce" else 1, x[1]))

    # ══════════════════════════════════════════════════════════════════════════
    # LAYER 1 TABLE — printed only if user said yes
    # ══════════════════════════════════════════════════════════════════════════
    if show_layer1:
        header("LAYER 1 — SAME-CANDLE SPIKE  (Delta ∧ Gamma both high simultaneously)", W)
        print(f"  {dim('Threshold: runtime percentile ≥')} {bold(str(STRENGTH_PCT))}th"
              f"  {dim('|  Delta min:')} {bold(str(DELTA_MIN_PCT))}%"
              f"  {dim('| Gamma min:')} {bold(str(GAMMA_MIN_PCT))}%\n")

        h1 = (
            f"  {'TIMESTAMP':<19}  {'STRIKE':>8}  {'SPOT':>8}  {'SIDE':>4}  "
            f"{'SCORE':>5}  {'STRENGTH':>18}  "
            f"{'DELTA':>10} {'D CHG%':>9}  {'GAMMA':>10} {'G CHG%':>9}  "
            f"{'VOLUME':>10} {'V%':>9}  {'PRICE':>8} {'P%':>9}  {'IV':>7}  "
            f"{'MAX BEST MOVE':>13}"   # ← NEW header column
        )
        print(bold(h1))
        sep(W)

        for ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv, max_move in layer1_rows:
            sc     = f"{score:5.1f}"
            bar    = strength_bar(score)
            lbl    = strength_label(score)
            side_c = bright_green("CE") if side == "ce" else bright_red("PE")

            # Max best move: green if positive, dim if zero/negative
            if max_move > 0:
                move_txt = bright_green(f"+{max_move:>6.2f} pts")
            elif max_move < 0:
                move_txt = red(f"{max_move:>7.2f} pts")
            else:
                move_txt = dim(f"  0.00 pts")

            print(
                f"  {dim(str(ts)[:19])}  "
                f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
                f"{bold(sc)}  {bar} {lbl}  "
                f"{rjust(dim(f'{dv:>10.4f}'), 10)} {rjust(color_pct(dp), 9)}  "
                f"{rjust(dim(f'{gv:>10.6f}'), 10)} {rjust(color_pct(gp), 9)}  "
                f"{rjust(dim(f'{vv:>10,.0f}'), 10)} {rjust(color_pct(vp, 50), 9)}  "
                f"{dim(f'{pv:>8.2f}')} {rjust(color_pct(pp), 9)}  "
                f"{dim(f'{iv:>7.2f}%')}  "
                f"{rjust(move_txt, 13)}"
            )

        if not layer1_rows:
            print(f"  {dim('No same-candle spikes found.')}")
        sep(W)
    else:
        print(dim(f"  [Layer 1 table skipped — {len(layer1_rows)} spikes found, included in XLSX]\n"))

    # ══════════════════════════════════════════════════════════════════════════
    # TOP ENTRIES
    # ══════════════════════════════════════════════════════════════════════════
    header("TOP ENTRIES — Delta + Gamma + Volume + Price increasing", W)

    top_rows = []
    skipped_entry_window = 0
    skipped_choppy       = 0

    for row in layer1_rows:
        ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv, max_move = row

        if ts.time() >= ENTRY_CUTOFF:
            continue
        if ts.time() < ENTRY_START:
            skipped_entry_window += 1
            continue
        if is_choppy_market(ts, spot_by_ts):
            skipped_choppy += 1
            continue

        # Use runtime DELTA_MIN_PCT and GAMMA_MIN_PCT (user-overridable)
        strong_checks = [
            dp > DELTA_MIN_PCT,
            gp > GAMMA_MIN_PCT,
            vp > VOLUME_MIN_PCT,
            pp > PRICE_MIN_PCT
        ]

        if all(strong_checks):
            entry_score = (
                dp * 0.35 +
                gp * 0.25 +
                min(vp, 300) * 0.25 +
                pp * 0.15
            )

            if SCR_ABOVE is not None and entry_score <= SCR_ABOVE:
                continue

            entry_price = pv
            exit_ts, exit_price, _, exit_reason = find_exit_after_entry(
                full, ts, stk, side, entry_price, col_map
            )
            pnl_points = exit_price - entry_price if exit_price is not None else np.nan

            top_rows.append((
                ts, stk, spot, side,
                entry_price, exit_ts, exit_price, pnl_points,
                max_move, entry_score, score, dp, gp, vp, pp,
                exit_reason,
                f"{sum(strong_checks)}/4 D,G,V,P strong"
            ))

    print(f"  {dim('Skipped (before 9:20 AM):')} {bold(str(skipped_entry_window))}   "
          f"{dim('Skipped (choppy market):')} {bold(str(skipped_choppy))}\n")

    top_rows.sort(key=lambda x: x[9], reverse=True)

    unique_rows = []
    seen_ts = set()
    for r in top_rows:
        ts_key = str(r[0])[:19]
        if ts_key in seen_ts:
            continue
        seen_ts.add(ts_key)
        unique_rows.append(r)

    h_top = (
        f"  {'ENTRY TIME':<19}  {'EXIT TIME':<19}  {'STRIKE':>8}  {'SPOT':>8}  {'SIDE':>4}  "
        f"{'ENTRY':>8}  {'EXIT':>8}  {'PNL':>8}  {'BEST PNL':>8}  "
        f"{'ENTRY SCR':>9}  {'SYS SCR':>7}  "
        f"{'D%':>9}  {'G%':>9}  {'V%':>9}  {'P%':>9}  {'EXIT REASON':<28}"
    )
    print(bold(h_top))
    sep(W)

    for ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move_pnl, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason in unique_rows[:TOP_N]:
        side_c  = bright_green("CE") if side == "ce" else bright_red("PE")
        pnl_txt = bright_green(f"{pnl_points:>8.2f}") if pnl_points >= 0 else bright_red(f"{pnl_points:>8.2f}")
        best_txt = bright_green(f"{best_move_pnl:>8.2f}") if best_move_pnl >= 0 else bright_red(f"{best_move_pnl:>8.2f}")
        print(
            f"  {dim(str(ts)[:19])}  {dim(str(exit_ts)[:19])}  "
            f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
            f"{entry_price:>8.2f}  {exit_price:>8.2f}  {pnl_txt}  {best_txt}  "
            f"{bold(f'{entry_score:>9.1f}')}  {bold(f'{sys_score:>7.1f}')}  "
            f"{rjust(color_pct(dp), 9)}  {rjust(color_pct(gp), 9)}  "
            f"{rjust(color_pct(vp, 50), 9)}  {rjust(color_pct(pp), 9)}  "
            f"{exit_reason:<28}"
        )

    if not unique_rows:
        print(f"  {dim('No top entries found.')}")
    sep(W)

    # ══════════════════════════════════════════════════════════════════════════
    # FINAL SELECTED TRADES (day rules applied)
    # ══════════════════════════════════════════════════════════════════════════
    header("FINAL SELECTED TRADES — Day Rules Applied (max "
           f"{MAX_ENTRIES_PER_DAY}/day, no overlapping positions, min "
           f"{MIN_GAP_BETWEEN_ENTRIES_MIN}-min gap after exit, no chop, "
           "9:20–15:00 entries, 3:30 PM exit)", W)

    chronological_candidates = sorted(unique_rows, key=lambda r: r[0])
    selected_trades = []
    last_exit_time  = None

    for row in chronological_candidates:
        ts      = row[0]
        exit_ts = row[5]

        if len(selected_trades) >= MAX_ENTRIES_PER_DAY:
            break
        if last_exit_time is not None:
            if ts <= last_exit_time:
                continue
            gap_min = (ts - last_exit_time).total_seconds() / 60
            if gap_min < MIN_GAP_BETWEEN_ENTRIES_MIN:
                continue

        selected_trades.append(row)
        last_exit_time = exit_ts

    print(bold(h_top))
    sep(W)

    total_pnl = 0.0
    for ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move_pnl, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason in selected_trades:
        side_c  = bright_green("CE") if side == "ce" else bright_red("PE")
        pnl_txt = bright_green(f"{pnl_points:>8.2f}") if pnl_points >= 0 else bright_red(f"{pnl_points:>8.2f}")
        best_txt = bright_green(f"{best_move_pnl:>8.2f}") if best_move_pnl >= 0 else bright_red(f"{best_move_pnl:>8.2f}")
        total_pnl += pnl_points
        print(
            f"  {dim(str(ts)[:19])}  {dim(str(exit_ts)[:19])}  "
            f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
            f"{entry_price:>8.2f}  {exit_price:>8.2f}  {pnl_txt}  {best_txt}  "
            f"{bold(f'{entry_score:>9.1f}')}  {bold(f'{sys_score:>7.1f}')}  "
            f"{rjust(color_pct(dp), 9)}  {rjust(color_pct(gp), 9)}  "
            f"{rjust(color_pct(vp, 50), 9)}  {rjust(color_pct(pp), 9)}  "
            f"{exit_reason:<28}"
        )

    if not selected_trades:
        print(f"  {dim('No trades survived the day-trading rules.')}")
    else:
        pnl_txt = bright_green(f"{total_pnl:.2f}") if total_pnl >= 0 else bright_red(f"{total_pnl:.2f}")
        print(f"\n  {dim('Total PnL (selected trades):')} {bold(pnl_txt)}")
    sep(W)

    # ══════════════════════════════════════════════════════════════════════════
    # SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    header("SUMMARY", W)
    print(f"  {dim('Session window              :')} {bold(SESSION_START.strftime('%H:%M'))} – {bold(SESSION_END.strftime('%H:%M'))}")
    print(f"  {dim('Thresholds used             :')} "
          f"STRENGTH≥{bold(str(STRENGTH_PCT))}th  WEAK≤{bold(str(WEAK_PCT))}th  "
          f"ΔΔ%>{bold(str(DELTA_MIN_PCT))}  ΓΓ%>{bold(str(GAMMA_MIN_PCT))}")
    print(f"  {dim('Layer 1 — Same-candle spikes:')} {bold(str(len(layer1_rows)))}")
    print(f"  {dim('Layer 1 table shown         :')} {bold('YES' if show_layer1 else 'NO')}")
    print(f"  {dim('Entry filters              :')} "
          f"VOL%>{bold(str(VOLUME_MIN_PCT))}  "
          f"PRICE%>{bold(str(PRICE_MIN_PCT))}  "
          f"SCR>{bold(str(SCR_ABOVE) if SCR_ABOVE is not None else 'OFF')}")
    print(f"  {dim('Top entries shown           :')} {bold(str(min(len(unique_rows), TOP_N)))}")
    print(f"  {dim('Skipped — before 9:20 AM    :')} {bold(str(skipped_entry_window))}")
    print(f"  {dim('Skipped — choppy market     :')} {bold(str(skipped_choppy))}")
    print(f"  {dim('Final selected trades       :')} {bold(str(len(selected_trades)))} / {MAX_ENTRIES_PER_DAY} max")
    print()

    # ── Save outputs ──────────────────────────────────────────────────────────
    save_cols = ["timestamp","strike","spot"]
    for side in sides:
        cm = col_map[side]
        save_cols += [
            cm["delta"], f"{side}_d_pct", f"{side}_d_xpct", f"{side}_d_rank", f"{side}_d_trend",
            cm["gamma"], f"{side}_g_pct", f"{side}_g_xpct", f"{side}_g_rank", f"{side}_g_trend",
            cm["volume"],f"{side}_v_pct", f"{side}_v_xpct", f"{side}_v_rank", f"{side}_v_trend",
            cm["price"], f"{side}_p_pct", f"{side}_p_xpct", f"{side}_p_rank", f"{side}_p_trend",
            cm["iv"],    f"{side}_score", f"{side}_same_spike",
            f"{side}_cross_bull", f"{side}_cross_bear",
        ]
    if SAVE_OUTPUTS:
        full[[c for c in save_cols if c in full.columns]].to_csv(OUT_CSV, index=False)

        if selected_trades:
            sel_df = pd.DataFrame(selected_trades, columns=[
                "entry_time","strike","spot","side","entry_price","exit_time","exit_price",
                "pnl_points","best_move_pnl","entry_score","sys_score","d_pct","g_pct","v_pct","p_pct",
                "exit_reason","entry_rule"
            ])
            sel_df.to_csv(OUT_CSV_SELECTED, index=False)

        _export_xlsx(full, layer1_rows, sides, col_map, save_cols, timestamps,
                     selected_trades=selected_trades)

    trade_details = []
    for ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move_pnl, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason in selected_trades:
        best_info = compute_best_move_info(full, ts, stk, side, entry_price, col_map)
        best_time = best_info["best_move_time"]
        missed_after_exit = (
            pd.notna(best_time) and
            pd.notna(exit_ts) and
            pd.Timestamp(best_time) > pd.Timestamp(exit_ts)
        )
        trade_details.append({
            "entry_time": str(ts)[:19],
            "exit_time": str(exit_ts)[:19],
            "best_move_time": str(best_time)[:19] if pd.notna(best_time) else "",
            "strike": float(stk),
            "side": side.upper(),
            "entry_price": round(float(entry_price), 2),
            "exit_price": round(float(exit_price), 2),
            "pnl": round(float(pnl_points), 2),
            "best_move_pnl": round(float(best_info["best_move_pnl"]), 2),
            "best_move_price": round(float(best_info["best_move_price"]), 2),
            "missed_after_exit": "YES" if missed_after_exit else "NO",
            "entry_score": round(float(entry_score), 1),
            "sys_score": round(float(sys_score), 1),
            "d_pct": round(float(dp), 2),
            "g_pct": round(float(gp), 2),
            "v_pct": round(float(vp), 2),
            "exit_reason": exit_reason,
        })

    return {
        "csv": str(CSV_PATH),
        "date": str(ANALYSIS_DATE),
        "trades": len(selected_trades),
        "pnl": round(float(total_pnl), 2),
        "best_move_pnl": round(float(sum(r[8] for r in selected_trades)), 2) if selected_trades else 0.0,
        "wins": sum(1 for r in selected_trades if r[7] >= 0),
        "layer1": len(layer1_rows),
        "top_entries": len(unique_rows),
        "trade_details": trade_details,
    }


def yes(value):
    return str(value).strip().lower() in ("y", "yes", "true", "1")


def convert_db_to_csv(db_path, overwrite=False):
    db = Path(db_path)
    if not db.exists() or db.stat().st_size == 0:
        return None
    csv_path = db.with_suffix(".csv")
    if csv_path.exists() and not overwrite:
        return csv_path
    with sqlite3.connect(db) as conn:
        df = pd.read_sql("SELECT * FROM oi_snapshots ORDER BY timestamp, strike", conn)
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"])
    df.to_csv(csv_path, index=False)
    return csv_path


def convert_db_folder(db_dir, overwrite=False):
    db_files = sorted(Path(db_dir).glob("oi_*.db"))
    converted = 0
    existing = 0
    empty = 0
    failed = 0
    for db in db_files:
        try:
            if db.stat().st_size == 0:
                empty += 1
                continue
            csv_path = db.with_suffix(".csv")
            already_exists = csv_path.exists() and not overwrite
            result = convert_db_to_csv(db, overwrite=overwrite)
            if result is None:
                empty += 1
            elif already_exists:
                existing += 1
            else:
                converted += 1
        except Exception as exc:
            failed += 1
            print(yellow(f"  DB convert failed: {db.name}: {exc}"))
    print(f"  DB to CSV: converted={converted} existing={existing} empty={empty} failed={failed}")


def date_from_oi_filename(path):
    match = re.search(r"oi_(\d{4})_(\d{2})_(\d{2})", Path(path).stem)
    if not match:
        return None
    yyyy, mm, dd = match.groups()
    return f"{yyyy}-{mm}-{dd}"


def parse_month(value):
    if value is None:
        return None
    raw = str(value).strip().lower()
    names = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    if raw in names:
        return names[raw]
    try:
        month = int(raw)
    except ValueError:
        raise ValueError(f"Invalid month: {value}")
    if not 1 <= month <= 12:
        raise ValueError(f"Month must be 1-12: {value}")
    return month


def batch_date_bounds(args):
    start = pd.to_datetime(args.from_date).date() if args.from_date else None
    end = pd.to_datetime(args.to_date).date() if args.to_date else None

    if args.year is not None:
        from_month = parse_month(args.from_month) or 1
        to_month = parse_month(args.to_month) or 12
        if from_month > to_month:
            raise ValueError("--from-month cannot be after --to-month for the same --year")
        month_start = pd.Timestamp(args.year, from_month, 1).date()
        month_end = (pd.Timestamp(args.year, to_month, 1) + pd.offsets.MonthEnd(0)).date()
        start = max(start, month_start) if start else month_start
        end = min(end, month_end) if end else month_end

    return start, end


def in_batch_date_range(file_date, start, end):
    dt = pd.to_datetime(file_date).date()
    if start and dt < start:
        return False
    if end and dt > end:
        return False
    return True


def print_group_summary(results, key_name):
    if not results:
        return
    df = pd.DataFrame(results)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if key_name == "month":
        df["period"] = df["date"].dt.strftime("%Y-%m")
        title = "MONTHLY SUMMARY"
    else:
        df["period"] = df["date"].dt.strftime("%Y")
        title = "YEARLY SUMMARY"

    grouped = (
        df.groupby("period", dropna=True)
          .agg(files=("csv", "count"),
               trades=("trades", "sum"),
               wins=("wins", "sum"),
               pnl=("pnl", "sum"),
               best_move_pnl=("best_move_pnl", "sum"))
          .reset_index()
    )
    grouped["win_rate"] = np.where(
        grouped["trades"] > 0,
        (grouped["wins"] / grouped["trades"] * 100).round(1),
        0.0,
    )

    header(title, 110)
    print(f"  {'PERIOD':<10} {'FILES':>6} {'TRADES':>7} {'WINS':>6} {'WIN%':>7} {'PNL':>10} {'BEST MOVE PNL':>14}")
    sep(110)
    for _, r in grouped.iterrows():
        pnl_txt = bright_green(f"{r['pnl']:>10.2f}") if r["pnl"] >= 0 else bright_red(f"{r['pnl']:>10.2f}")
        best_txt = bright_green(f"{r['best_move_pnl']:>14.2f}") if r["best_move_pnl"] >= 0 else bright_red(f"{r['best_move_pnl']:>14.2f}")
        print(f"  {r['period']:<10} {int(r['files']):>6} {int(r['trades']):>7} {int(r['wins']):>6} "
              f"{r['win_rate']:>6.1f}% {pnl_txt} {best_txt}")
    sep(110)


def run_batch(args, show_layer1):
    global CSV_PATH, ANALYSIS_DATE, SAVE_OUTPUTS

    if args.convert_db:
        convert_db_folder(args.db_dir, overwrite=args.overwrite_csv)

    csv_files = sorted(Path(args.csv_dir).glob("oi_*.csv"))
    if not csv_files:
        print(red(f"No oi_*.csv files found in {args.csv_dir}"))
        return []

    start_date, end_date = batch_date_bounds(args)
    if start_date or end_date:
        print(f"  Batch date filter: {start_date or 'START'} to {end_date or 'END'}")

    old_save_outputs = SAVE_OUTPUTS
    SAVE_OUTPUTS = False
    results = []
    printed_table_header = False
    try:
        for csv_file in csv_files:
            file_date = date_from_oi_filename(csv_file)
            if not file_date:
                continue
            if not in_batch_date_range(file_date, start_date, end_date):
                continue
            CSV_PATH = str(csv_file)
            ANALYSIS_DATE = file_date
            try:
                with contextlib.redirect_stdout(io.StringIO()):
                    result = run_single_analysis(show_layer1=False)
                results.append(result)
                if not printed_table_header:
                    print()
                    print(f"  {'CSV FILE':<18} {'DATE':<10} {'TRADES':>6} {'PNL':>10} {'BEST MOVE PNL':>14}")
                    sep(72)
                    printed_table_header = True
                pnl_txt = bright_green(f"{result['pnl']:>8.2f}") if result["pnl"] >= 0 else bright_red(f"{result['pnl']:>8.2f}")
                best_txt = bright_green(f"{result['best_move_pnl']:>14.2f}") if result["best_move_pnl"] >= 0 else bright_red(f"{result['best_move_pnl']:>14.2f}")
                print(f"  {csv_file.name:<18} {ANALYSIS_DATE:<10} "
                      f"{result['trades']:>6} {pnl_txt} {best_txt}")
                if result["trades"] > 0:
                    print(f"    {'ENTRY TIME':<19} {'EXIT TIME':<19} {'BEST TIME':<19} {'SIDE':<4} {'STRIKE':>8} "
                          f"{'PNL':>10} {'BEST MOVE PNL':>14} {'BEST PX':>8} {'MISSED':>6} "
                          f"{'ENTRY SCR':>9} {'SYS SCR':>7} {'D%':>8} {'G%':>8} {'V%':>8}  {'EXIT REASON'}")
                    for trade in result["trade_details"]:
                        trade_pnl = bright_green(f"{trade['pnl']:>10.2f}") if trade["pnl"] >= 0 else bright_red(f"{trade['pnl']:>10.2f}")
                        trade_best = bright_green(f"{trade['best_move_pnl']:>14.2f}") if trade["best_move_pnl"] >= 0 else bright_red(f"{trade['best_move_pnl']:>14.2f}")
                        missed_txt = bright_red(f"{trade['missed_after_exit']:>6}") if trade["missed_after_exit"] == "YES" else dim(f"{trade['missed_after_exit']:>6}")
                        print(f"    {trade['entry_time']:<19} {trade['exit_time']:<19} {trade['best_move_time']:<19} "
                              f"{trade['side']:<4} {trade['strike']:>8.1f} {trade_pnl} {trade_best} "
                              f"{trade['best_move_price']:>8.2f} {missed_txt} "
                              f"{trade['entry_score']:>9.1f} {trade['sys_score']:>7.1f} "
                              f"{trade['d_pct']:>7.2f}% {trade['g_pct']:>7.2f}% {trade['v_pct']:>7.2f}%  "
                              f"{trade['exit_reason']}")
                    print()
            except Exception as exc:
                print(yellow(f"  skipped {csv_file.name}: {exc}"))
    finally:
        SAVE_OUTPUTS = old_save_outputs

    total_pnl = sum(r["pnl"] for r in results)
    total_best_move_pnl = sum(r["best_move_pnl"] for r in results)
    total_trades = sum(r["trades"] for r in results)
    total_wins = sum(r["wins"] for r in results)
    win_rate = (total_wins / total_trades * 100) if total_trades else 0.0
    header("BATCH SUMMARY", 110)
    print(f"  Files analysed : {len(results)}")
    print(f"  Trades         : {total_trades}")
    print(f"  Wins           : {total_wins} ({win_rate:.1f}%)")
    print(f"  Total PnL      : {total_pnl:.2f}")
    print(f"  Best Move PnL  : {total_best_move_pnl:.2f}")
    print(f"  Note           : PnL is booked at strategy exit; Best Move PnL is hindsight max after entry until session end.")
    print(f"                   If Best Move is much higher, the exit rule closed before the later move.")
    print()

    if yes(args.monthly):
        print_group_summary(results, "month")
    if yes(args.yearly):
        print_group_summary(results, "year")
    return results


def main():
    apply_run_mode_defaults()
    args = parse_cli_args()
    apply_cli_args(args)
    if args.interactive or not has_cli_overrides(args):
        show_layer1 = ask_user_options()
    else:
        show_layer1 = args.show_layer1

    if args.all_csv or args.convert_db:
        run_batch(args, show_layer1)
    else:
        run_single_analysis(show_layer1)


if __name__ == "__main__":
    main()
