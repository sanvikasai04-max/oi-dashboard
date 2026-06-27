r"""
Script 3: Options Flow Strength Analyser

Default compile check:
  python -m py_compile .\3_strength_analyser.py

Default single-day command:
  python .\3_strength_analyser.py --csv oi_2024_12_31.csv --date 2024-12-31 --console-xlsx

1-minute console Excel:
  python .\3_strength_analyser.py --csv oi_2024_12_31.csv --date 2024-12-31 --print-interval 1min --console-xlsx

5-minute console Excel:
  python .\3_strength_analyser.py --csv oi_2024_12_31.csv --date 2024-12-31 --print-interval 5min --console-xlsx

Current entry/exit model
========================

Multi-factor cross-strike + cross-candle confirmation engine.

Three pillars scored independently per snapshot:

  Volume pillar (weight 20%)
  --------------------------
  Minimum own-side volume pct: >= 80 (STRICT_MIN_VOLUME_PCT)
  Bonus tier counts per strike:
    >= 80%  = base confirmation
    >= 100% = V100 bonus
    >= 150% = V150 bonus
    >= 200% = V200 bonus
  More strikes hitting higher tiers = stronger score.
  Opposite side volume: either sell pressure (>= 80) or exits (< 0) both valid.

  Delta pillar (weight 40%)
  -------------------------
  Minimum own-side delta pct: >= 2 (STRICT_MIN_DELTA_PCT)
  Bonus tiers: >=3, >=4, >=5 (counted across nearby strikes).
  Opposite side: delta must DROP (<= -2, -4, -6); more strikes = stronger score.

  Price pillar (weight 40%)
  -------------------------
  Minimum own-side price pct: >= 3 (STRICT_MIN_PRICE_PCT)
  Bonus tiers: >=4, >=5, >=6, >=7.
  Opposite side: price must FALL (<= -3, -5).

Cross-strike confirmation:
  Own side: at least STRICT_MIN_SAME_STRIKES (default 3) nearby strikes must pass.
  Opposite side: at least STRICT_MIN_OPP_STRIKES (default 2) must confirm weakness.

Volume-neutral partial confirmation:
  If volume is flat (>=0) but delta+price are strong, snapshot still counts.
  This handles the rule: "delta spike + price spike but volume neutral = still confirm".
  Partial confirmation reduces effective weight (not counted at full volume bonus).

Cross-candle confirmation (ENTRY):
  Window: last STRICT_CONFIRM_WINDOW snapshots (default 5).
  Minimum: STRICT_MIN_CONFIRM_BARS (default 3) must pass the above rules.
  Non-consecutive confirmations are allowed (e.g., candles 1, 3, 4 out of 5).
  Growing strength: the LAST confirmation score must be >= the FIRST.
  Example: candles 1, 2, 4 strong out of 5 => 3 of 5 confirmed, entry ok if growing.

EXIT criteria:
  1. Fixed stop loss of STOP_LOSS_PTS (default 30) option points from entry.
  2. Opposite-side mirror confirmation:
     - Wait for STRICT_EXIT_CONFIRM_BARS (default 2) opposite confirmations.
     - Latest opposite confirmation must be >= first (growing exit pressure).
     - Early exit if 2nd opposite confirmation is > 1.5x score of 1st (strong surge).
  3. Force exit at FORCE_EXIT (default 15:25).

Tunable parameters:
  --strict-volume-pct 80        Own-side minimum volume pct
  --strict-delta-pct 2          Own-side minimum delta pct (opposite must drop by same)
  --strict-price-pct 3          Own-side minimum price pct (opposite must drop by same)
  --strict-same-strikes 3       Minimum own-side confirming strikes
  --strict-opp-strikes 2        Minimum opposite-side confirming strikes
  --strict-window 5             Candle lookback window
  --strict-confirm-bars 3       Minimum confirmations inside window
  --strict-exit-bars 2          Opposite confirmations needed to exit
  --strict-require-volume       Require volume threshold (disables neutral-volume partial confirm)

Monthly mode:
  python .\3_strength_analyser.py --monthly --month jan --year 2024

  python 3_strength_analyser_june27.py --console-xlsx --strict-window 5 --strict-confirm-bars 2 --strict-exit-bars 2 --exit-surge-ratio 1.5

  #--spike-override-score 80  -> from 1st strong candle 


  
"""
import argparse
import contextlib
import io
import re
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import time as dtime

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

# USER CONFIG
CSV_PATH         = "oi_2026_06_09.csv"
ANALYSIS_DATE    = "2026-06-03"
SIDE             = "both"          # "ce" | "pe" | "both"
OI_DATA_MODE     = "back"          # "back" | "live"
BACK_OI_DATA_DIR = Path(__file__).resolve().parent
LIVE_OI_DATA_DIR = Path(__file__).resolve().parents[3] / "StrategyBuilder" / "Strategy3MAGreeks" / "live_Saidata_weeklyexp"
OI_PRINT_INTERVAL = "1min"         # "5min" default; pass "1min" to print every candle
CROSS_CANDLES    = 5               # lookback window for cross-candle trend
STRIKES_NEARBY   = 2
ATM_RANGE_POINTS  = 500
STRIKE_STEP       = 50
STRENGTH_PCT     = 60              # percentile threshold for "strong" (runtime)
WEAK_PCT         = 40              # percentile threshold for "weak"  (runtime)
OI_TABLE_PRICE_PCT = 3             # also show rows where option price builds >= 3%
OI_TABLE_VOLUME_PCT = 3            # and option volume builds >= 3%
OI_TABLE_PRICE_DROP_PCT = -3       # also show opposite-side weakness when price falls <= -3%

SESSION_START    = dtime(9, 20)    # filter: keep rows from this time onward
SESSION_END      = dtime(15, 30)   # filter: keep rows up to this time
OUT_CSV          = "strength_report.csv"
OUT_XLSX         = "strength_report.xlsx"
REPORTS_DIR      = Path(__file__).resolve().parent / "reports"
OUT_CONSOLE_XLSX = str(REPORTS_DIR / "console_output.xlsx")
TOP_N = 20
ALLOW_OVERLAPPING_TRADES = False  # False = one active trade only; next entry after exit
ALLOW_SIDE_SWITCH = True          # True = if opposite side fires while trade active, close and flip

# Entry/exit model: multi-factor cross-strike + cross-candle confirmation.
#
# THREE PILLARS per snapshot (delta 40%, price 40%, volume 20%):
#   Volume: own >= 80%; bonus tiers >= 100/150/200%; opposite sell or exit.
#   Delta:  own >= 2%; bonus tiers >= 3/4/5%; opposite drops <= -2/-4/-6%.
#   Price:  own >= 3%; bonus tiers >= 4/5/6/7%; opposite drops <= -3/-5%.
#
# Cross-strike: own >= STRICT_MIN_SAME_STRIKES; opposite >= STRICT_MIN_OPP_STRIKES.
# Volume-neutral: if delta+price strong but volume flat, snapshot still counts.
#
# Cross-candle ENTRY: last 5 snapshots, at least 3 must confirm (non-consecutive
# allowed). Last confirmation score >= first (growing strength).
#
# EXIT: opposite side mirror; 2 opposite confirmations with growing strength.
# Early exit if 2nd opposite score > 1.5x first (strong surge).
# Fixed stop loss and force exit at FORCE_EXIT.
ENTRY_START = dtime(9, 20)
USE_STRICT_CROSS_ENTRY = True
STRICT_MIN_VOLUME_PCT = 80
STRICT_MIN_DELTA_PCT = 2
STRICT_MIN_PRICE_PCT = 3
STRICT_MIN_SAME_STRIKES = 3
STRICT_MIN_OPP_STRIKES = 2
STRICT_CONFIRM_WINDOW = 5
STRICT_MIN_CONFIRM_BARS = 3
STRICT_ALLOW_VOLUME_NEUTRAL = True
STRICT_EXIT_CONFIRM_BARS = 2
STRICT_EXIT_SURGE_RATIO  = 1.5   # early exit if 2nd opp score > this × 1st score
# Spike override: if the current snapshot composite score is >= this value AND
# at least 1 prior confirmation exists in the window, enter immediately without
# waiting for STRICT_MIN_CONFIRM_BARS. Captures explosive moves like 12:32 where
# volume 200-800%, delta 100%+, price 5-6% all appear together for the first time.
# Set to 0 to disable (always require full confirm bars).
STRICT_SPIKE_OVERRIDE_SCORE = 60
_STRICT_BEST_CACHE = {}

# Exit config: fixed stop, strict opposite confirmation, or force exit only.
STOP_LOSS_PTS = 30

ENTRY_CUTOFF = dtime(15, 0)   # no new entries after 3 PM
FORCE_EXIT   = dtime(15, 25)  # force exit near market close, after late moves can mature

# Cross-side OI buildup confirmation:
# Bullish setup = CE buying across nearby strikes + PE selling confirmation.
# Bearish setup = PE buying across nearby strikes + CE selling confirmation.
# This is based on the daily OI buildup examples:
#   CE: delta up, volume up, price up.
#   PE opposite: delta weakens, volume up, price down.
# Gamma is a score bonus only, because gamma can expand on both sides.
OI_BUILD_MIN_BUY_D_PCT = 6
OI_BUILD_MIN_BUY_G_PCT = 2
OI_BUILD_MIN_BUY_V_PCT = 40
OI_BUILD_MIN_BUY_P_PCT = 8
OI_BUILD_MIN_SELL_D_DROP_PCT = 3
OI_BUILD_MIN_SELL_V_PCT = 40
OI_BUILD_MIN_SELL_P_DROP_PCT = 5
OI_BUILD_MIN_BUY_STRIKES = 2
OI_BUILD_MIN_SELL_STRIKES = 2
OI_BUILD_STRONG_BONUS = 4
OI_BUILD_GAMMA_BONUS_PER_STRIKE = 1

# ANSI helpers
def _c(t, code): return f"\033[{code}m{t}\033[0m"
def green(t):        return _c(t, "32")
def bright_green(t): return _c(t, "92")
def red(t):          return _c(t, "31")
def bright_red(t):   return _c(t, "91")
def yellow(t):       return _c(t, "33")
def orange(t):       return _c(t, "38;5;208")
def cyan(t):         return _c(t, "36")
def magenta(t):      return _c(t, "35")
def dim(t):          return _c(t, "2")
def bold(t):         return _c(t, "1")

def strip_ansi(s):
    return re.sub(r'\033\[[0-9;]*m', '', str(s))

def rjust(s, w):
    return ' ' * max(0, w - len(strip_ansi(s))) + s

def ljust(s, w):
    return s + ' ' * max(0, w - len(strip_ansi(s)))

def color_pct(val, extreme=20.0):
    if pd.isna(val): return dim("    n/a ")
    sign = "^" if val > 0 else "v" if val < 0 else " "
    txt  = f"{sign}{abs(val):6.2f}%"
    if val > extreme:    return bright_green(txt)
    elif val > 0:        return green(txt)
    elif val < -extreme: return bright_red(txt)
    elif val < 0:        return red(txt)
    return dim(txt)

def strength_bar(score, width=10):
    if score is None or pd.isna(score):
        return dim("[n/a]".ljust(width + 2))

    filled = int(round(float(score) / 100 * width))
    filled = max(0, min(width, filled))

    bar = "#" * filled + "-" * (width - filled)
    txt = f"[{bar}]"

    if score >= 75: return bright_green(txt)
    if score >= 50: return green(txt)
    if score >= 25: return yellow(txt)
    return red(txt)

def strength_label(score):
    if score >= 75: return bright_green("STRONG  ")
    if score >= 50: return green("MODERATE")
    if score >= 25: return yellow("WEAK    ")
    return red("VERY WK ")

def sep(w=130): print(dim("-" * w))
def header(title, w=130):
    print(bold("=" * w))
    print(bold(f"  {title}"))
    print(bold("=" * w))

class TeeConsole:
    def __init__(self, wrapped):
        self.wrapped = wrapped
        self.parts = []

    def write(self, text):
        self.parts.append(text)
        return self.wrapped.write(text)

    def flush(self):
        return self.wrapped.flush()

    def isatty(self):
        return getattr(self.wrapped, "isatty", lambda: False)()

    def getvalue(self):
        return "".join(self.parts)


def ansi_to_rich_text(line):
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    color_map = {
        "31": "FF4444",
        "32": "00CC00",
        "33": "FFFF00",
        "35": "FF66FF",
        "36": "00B0F0",
        "91": "FF5555",
        "92": "00FF00",
        "38;5;208": "FFA500",
    }

    result = CellRichText()
    pos = 0
    color = "C0C0C0"
    bold_on = False
    dim_on = False

    for match in re.finditer(r"\033\[([0-9;]*)m", line):
        text = line[pos:match.start()]
        if text:
            fg = "777777" if dim_on else color
            result.append(TextBlock(InlineFont(rFont="Consolas", sz=9, color=fg, b=bold_on), text))

        code = match.group(1) or "0"
        if code == "0":
            color = "C0C0C0"
            bold_on = False
            dim_on = False
        elif code == "1":
            bold_on = True
        elif code == "2":
            dim_on = True
        elif code in color_map:
            color = color_map[code]
            dim_on = False
        elif code.startswith("38;5;"):
            color = color_map.get(code, "C0C0C0")
            dim_on = False
        pos = match.end()

    tail = line[pos:]
    if tail:
        fg = "777777" if dim_on else color
        result.append(TextBlock(InlineFont(rFont="Consolas", sz=9, color=fg, b=bold_on), tail))
    return result if result else strip_ansi(line)


def export_console_xlsx(console_text, output_path=OUT_CONSOLE_XLSX):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_path = Path(output_path)
    if not output_path.is_absolute() and output_path.parent == Path("."):
        output_path = REPORTS_DIR / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Console Output"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 230

    header_cell = ws["A1"]
    header_cell.value = "Console Output"
    header_cell.font = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    header_cell.fill = PatternFill("solid", fgColor="1F4E79")

    dark_fill = PatternFill("solid", fgColor="0B0B0B")
    for row_idx, line in enumerate(console_text.splitlines(), 2):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = ansi_to_rich_text(line)
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 15

    wb.save(output_path)
    print(f"Console Excel written: {output_path.resolve()}")

def parse_month(value):
    raw = str(value).strip().lower()
    names = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    if raw in names:
        return names[raw]
    month = int(raw)
    if month < 1 or month > 12:
        raise ValueError("month must be 1-12 or month name like jan")
    return month

def parse_time_arg(value):
    try:
        hour, minute = str(value).strip().split(":", 1)
        return dtime(int(hour), int(minute))
    except Exception as exc:
        raise argparse.ArgumentTypeError(f"Invalid time '{value}'. Use HH:MM, e.g. 09:25") from exc

def parse_print_interval_arg(value):
    raw = str(value).strip().lower()
    aliases = {
        "1": "1min",
        "1m": "1min",
        "1min": "1min",
        "1mins": "1min",
        "5": "5min",
        "5m": "5min",
        "5min": "5min",
        "5mins": "5min",
    }
    if raw not in aliases:
        raise argparse.ArgumentTypeError("Invalid print interval. Use 1min or 5min.")
    return aliases[raw]

def print_interval_minutes():
    return 1 if OI_PRINT_INTERVAL == "1min" else 5

def is_print_interval_timestamp(ts):
    if print_interval_minutes() == 1:
        return True
    ts = pd.Timestamp(ts)
    return ts.minute % print_interval_minutes() == 0

def month_name(month):
    return pd.Timestamp(2000, month, 1).strftime("%B")

def date_from_oi_filename(path):
    match = re.search(r"oi_(\d{4})_(\d{2})_(\d{2})", Path(path).stem)
    if not match:
        return None
    yyyy, mm, dd = match.groups()
    return pd.Timestamp(int(yyyy), int(mm), int(dd)).date()

def monthly_date_range(year, month, include_next_week=True):
    start = pd.Timestamp(year, month, 1)
    end = start + pd.offsets.MonthEnd(0)
    if include_next_week:
        end = end + pd.Timedelta(days=7)
    return start.date(), end.date()

def trading_dates_in_csv(csv_file):
    try:
        dates = pd.read_csv(csv_file, usecols=["timestamp"], parse_dates=["timestamp"])
    except Exception:
        file_date = date_from_oi_filename(csv_file)
        return [file_date] if file_date else []

    if dates.empty:
        return []

    return sorted(dates["timestamp"].dt.date.dropna().unique())

def monthly_signal_date(csv_file):
    file_date = date_from_oi_filename(csv_file)
    dates = trading_dates_in_csv(csv_file)
    if not dates:
        return file_date
    if not file_date:
        return dates[-1]

    before_expiry = [d for d in dates if d < file_date]
    fridays = [d for d in before_expiry if pd.Timestamp(d).weekday() == 4]
    if fridays:
        return fridays[-1]
    if before_expiry:
        return before_expiry[-1]
    if file_date in dates:
        return file_date
    return dates[-1]

def find_exit_after_entry(
    full,
    entry_ts,
    strike,
    side,
    entry_price,
    col_map,
    snapshot_analysis=None,
    strict_entry=False,
):
    """
    Exit engine (mirror of entry engine).

    Exit rules (matching your strategy notes):

    1. Stop loss: fixed STOP_LOSS_PTS option points below entry.
    2. Opposite-side confirmation exit:
       - Watch opposite side using the same multi-factor snapshot check.
       - Wait for STRICT_EXIT_CONFIRM_BARS confirmations (default 2) in the
         last exit-window snapshots.
       - The LATEST opposite confirmation must be >= the FIRST opposite
         confirmation in score (growing exit strength).
       - If after the 1st opposite confirmation the next candle is STRONGER,
         exit immediately even without reaching the full bar count.
    3. Force exit at FORCE_EXIT time.
    """
    cm = col_map[side]
    stop_loss_pts = STOP_LOSS_PTS

    trade = full[
        (full["timestamp"] > entry_ts) &
        (full["strike"] == strike)
    ].copy().sort_values("timestamp").reset_index(drop=True)

    if trade.empty:
        return entry_ts, entry_price, 0, "No future data"

    opp = opposite_side(side)
    opp_confirm_history = []   # list of (ts, score, units) for opposite confirmations

    for i, r in trade.iterrows():
        price = r[cm["price"]]

        # Force exit
        if r["timestamp"].time() >= FORCE_EXIT:
            pnl = price - entry_price
            return r["timestamp"], price, pnl, f"Exit: force exit {FORCE_EXIT.strftime('%H:%M')}"

        pnl_now = price - entry_price

        # Stop loss
        if pnl_now <= -stop_loss_pts:
            return r["timestamp"], price, pnl_now, f"Exit: stop loss -{stop_loss_pts}"

        # Check opposite side
        opp_ok, opp_score, opp_own, opp_opp_count, opp_units, opp_strike, opp_reason = strict_best_snapshot(
            full, r["timestamp"], strike, opp
        )

        if opp_ok:
            opp_confirm_history.append({
                "ts": r["timestamp"],
                "score": opp_score,
                "units": opp_units,
            })

            # If we have enough opposite confirmations
            if len(opp_confirm_history) >= STRICT_EXIT_CONFIRM_BARS:
                first_opp = opp_confirm_history[0]
                last_opp  = opp_confirm_history[-1]

                # Growing exit strength: last >= first
                if last_opp["units"] >= first_opp["units"]:
                    return r["timestamp"], price, pnl_now, (
                        f"Exit: opposite {opp.upper()} confirm {opp_reason} "
                        f"STRIKE={opp_strike} CNT={len(opp_confirm_history)} "
                        f"FIRST_UNITS={first_opp['units']} LAST_UNITS={last_opp['units']}"
                    )

            # Early exit: 2nd confirmation is much stronger than 1st (score > STRICT_EXIT_SURGE_RATIO x first)
            if len(opp_confirm_history) >= 2:
                first_opp = opp_confirm_history[0]
                last_opp  = opp_confirm_history[-1]
                if last_opp["score"] > first_opp["score"] * STRICT_EXIT_SURGE_RATIO and last_opp["units"] >= first_opp["units"]:
                    return r["timestamp"], price, pnl_now, (
                        f"Exit: strong opposite surge {opp.upper()} "
                        f"STRIKE={opp_strike} CNT={len(opp_confirm_history)} "
                        f"SCORE_RATIO={last_opp['score']}/{first_opp['score']}"
                    )
        else:
            # Reset if opposite side weakens (no confirmation)
            # Keep history but don't grow it on misses - allows non-consecutive confirms
            # as per your rule: "out of 5, minimum 3 cross candle should have confirmation"
            # We don't reset here intentionally so non-consecutive confirmations count.
            pass

    last = trade.iloc[-1]
    exit_price = last[cm["price"]]
    pnl = exit_price - entry_price
    return last["timestamp"], exit_price, pnl, "Exit: session end"


def compute_strike_best_move(full, entry_ts, strike, side, entry_price, col_map):
    """
    Best move for the SAME STRIKE only.
    Checks future candles after entry for the same strike and same side price.
    """
    cm = col_map[side]
    future = full[
        (full["timestamp"] > entry_ts) &
        (full["strike"] == strike)
    ].copy().sort_values("timestamp")

    if future.empty:
        return 0.0, entry_ts, entry_price

    best_idx = future[cm["price"]].idxmax()
    best_row = future.loc[best_idx]
    best_price = float(best_row[cm["price"]])
    best_move = best_price - float(entry_price)

    return round(best_move, 2), best_row["timestamp"], round(best_price, 2)

def compute_strike_bad_move(full, entry_ts, strike, side, entry_price, col_map):
    """
    Bad move for the SAME STRIKE only.
    Shows the maximum adverse option-price move after entry.
    """
    cm = col_map[side]
    future = full[
        (full["timestamp"] > entry_ts) &
        (full["strike"] == strike)
    ].copy().sort_values("timestamp")

    if future.empty:
        return 0.0, entry_ts, entry_price

    low_idx = future[cm["price"]].astype(float).idxmin()
    low_row = future.loc[low_idx]
    low_price = float(low_row[cm["price"]])
    bad_move = low_price - float(entry_price)

    return round(bad_move, 2), low_row["timestamp"], round(low_price, 2)

def strict_cross_snapshot(full, ts, strike, side):
    """
    Multi-factor snapshot confirmation engine.

    Scores each snapshot using three separate pillars:
      1. Volume pct  - own >=80 required; bonus for >=100 / >=150 / >=200 pct strikes
      2. Delta pct   - own >=2; bonus for >=3/4/5; opposite must drop (<=-2)
      3. Price pct   - own >=3; bonus for >=4/5/6/7; opposite must fall (<=-3)

    Opposite side is considered valid if BOTH sides of each pillar agree:
      CE rising => PE price and delta must fall.
      PE volume may rise (sell pressure) OR fall (PE exits). Both are valid.

    Volume-neutral own-side:
      If volume is flat but delta+price are strong, the snapshot still counts
      (weighted less via STRICT_ALLOW_VOLUME_NEUTRAL). This matches the rule:
      delta spike + price spike but volume neutral = still count as confirmation.

    Returns: ok, composite_score, own_count, opp_count, reason
    """
    nearby = full[
        (full["timestamp"] == ts) &
        (full["strike"] >= strike - STRIKES_NEARBY * STRIKE_STEP) &
        (full["strike"] <= strike + STRIKES_NEARBY * STRIKE_STEP)
    ].copy()

    if nearby.empty:
        return False, 0, 0, 0, "STRICT_NO_NEARBY"

    own = "ce" if side == "ce" else "pe"
    opp = "pe" if side == "ce" else "ce"
    tag = "STRICT_BULL" if side == "ce" else "STRICT_BEAR"

    # OWN SIDE
    own_v = nearby[f"{own}_v_pct"]
    v80   = _count_true(own_v >= 80)
    v100  = _count_true(own_v >= 100)
    v150  = _count_true(own_v >= 150)
    v200  = _count_true(own_v >= 200)

    own_d = nearby[f"{own}_d_pct"]
    d2    = _count_true(own_d >= 2)
    d3    = _count_true(own_d >= 3)
    d4    = _count_true(own_d >= 4)
    d5    = _count_true(own_d >= 5)

    own_p = nearby[f"{own}_p_pct"]
    p3    = _count_true(own_p >= 3)
    p4    = _count_true(own_p >= 4)
    p5    = _count_true(own_p >= 5)
    p6    = _count_true(own_p >= 6)
    p7    = _count_true(own_p >= 7)

    # Per-strike full confirmation (all three pillars)
    own_full_confirm = (
        (own_d >= STRICT_MIN_DELTA_PCT) &
        (own_p >= STRICT_MIN_PRICE_PCT) &
        (own_v >= STRICT_MIN_VOLUME_PCT)
    )
    own_count = _count_true(own_full_confirm)

    # Volume-neutral partial confirmation (delta+price strong, volume not negative)
    if STRICT_ALLOW_VOLUME_NEUTRAL:
        own_partial = (
            (own_d >= STRICT_MIN_DELTA_PCT) &
            (own_p >= STRICT_MIN_PRICE_PCT) &
            (own_v >= 0)
        )
        own_count = max(own_count, _count_true(own_partial))

    # OPPOSITE SIDE
    opp_d = nearby[f"{opp}_d_pct"]
    opp_p = nearby[f"{opp}_p_pct"]
    opp_v = nearby[f"{opp}_v_pct"]

    opp_d2 = _count_true(opp_d <= -2)
    opp_d4 = _count_true(opp_d <= -4)
    opp_d6 = _count_true(opp_d <= -6)
    opp_p3 = _count_true(opp_p <= -3)
    opp_p5 = _count_true(opp_p <= -5)

    opp_confirm = (
        (opp_d <= -STRICT_MIN_DELTA_PCT) &
        (opp_p <= -STRICT_MIN_PRICE_PCT) &
        ((opp_v >= STRICT_MIN_VOLUME_PCT) | (opp_v < 0))
    )
    opp_count = _count_true(opp_confirm)

    # COMPOSITE SCORE: delta 40%, price 40%, volume 20%
    volume_score  = (v80 * 2) + (v100 * 1) + (v150 * 2) + (v200 * 3)
    delta_score   = (d2 * 3)  + (d3 * 2)   + (d4 * 2)   + (d5 * 3)
    price_score   = (p3 * 3)  + (p4 * 2)   + (p5 * 2)   + (p6 * 2)   + (p7 * 3)
    opp_score_pts = (opp_d2 * 2) + (opp_d4 * 2) + (opp_d6 * 2) + (opp_p3 * 2) + (opp_p5 * 2)

    score = int(
        delta_score * 0.40 +
        price_score * 0.40 +
        volume_score * 0.20 +
        opp_score_pts
    )

    ok = own_count >= STRICT_MIN_SAME_STRIKES and opp_count >= STRICT_MIN_OPP_STRIKES
    reason = (
        f"{tag} OWN={own_count} OPP={opp_count} "
        f"V80={v80} V100={v100} V150={v150} V200={v200} "
        f"D2={d2} D3={d3} D4={d4} D5={d5} "
        f"P3={p3} P4={p4} P5={p5} P6={p6} P7={p7} "
        f"OppD2={opp_d2} OppD4={opp_d4} OppP3={opp_p3} OppP5={opp_p5}"
    )
    return ok, score, own_count, opp_count, reason

def strict_candidate_strikes(full, ts, strike):
    nearby = full[
        (full["timestamp"] == ts) &
        (full["strike"] >= strike - STRIKES_NEARBY * STRIKE_STEP) &
        (full["strike"] <= strike + STRIKES_NEARBY * STRIKE_STEP)
    ]["strike"].dropna().unique()
    return sorted(nearby)

def strict_best_snapshot(full, ts, strike, side):
    cache_key = (id(full), pd.Timestamp(ts), float(strike), side)
    cached = _STRICT_BEST_CACHE.get(cache_key)
    if cached is not None:
        return cached

    best = None
    for candidate_strike in strict_candidate_strikes(full, ts, strike):
        ok, score, own_count, opp_count, reason = strict_cross_snapshot(
            full, ts, candidate_strike, side
        )
        strength_units = own_count + opp_count
        candidate = (
            ok,
            score,
            own_count,
            opp_count,
            strength_units,
            candidate_strike,
            reason,
        )
        if best is None:
            best = candidate
            continue
        if (ok, strength_units, score) > (best[0], best[4], best[1]):
            best = candidate

    if best is None:
        best = (False, 0, 0, 0, 0, strike, "STRICT_NO_NEARBY")
    _STRICT_BEST_CACHE[cache_key] = best
    return best

def strict_cross_entry_ok(full, ts, strike, side):
    """
    Cross-candle confirmation engine.

    Rules (matching your strategy notes):

    1. Look back at the last STRICT_CONFIRM_WINDOW snapshots (default 5).
    2. Count how many snapshots pass the multi-factor confirmation test.
       Require at least STRICT_MIN_CONFIRM_BARS (default 3) out of the window.
    3. Growing strength: the LAST valid confirmation snapshot must have a
       composite score >= the FIRST valid confirmation snapshot. If the move is
       fading, we wait.
    4. Partial candle: if a candle has delta+price strong but volume neutral,
       it still counts as a confirmation (at partial weight = 0.7 of a full
       confirmation). This matches the note: "only delta spike and price spike
       but volume neutral - still take strong confirmation".
    5. If after 5-10 candles another strong spike arrives with higher composite
       score than earlier, that is a strong hold/re-entry signal.

    Returns: ok, last_score, reason_string
    """
    if not USE_STRICT_CROSS_ENTRY:
        return False, 0, "STRICT_OFF"

    timestamps = [x for x in sorted(full["timestamp"].unique()) if x <= ts]
    recent = timestamps[-STRICT_CONFIRM_WINDOW:]

    # Collect all snapshot results in the window
    all_snaps = []
    for snap_ts in recent:
        ok, score, own_count, opp_count, strength_units, snap_strike, reason = strict_best_snapshot(
            full, snap_ts, strike, side
        )
        all_snaps.append({
            "ts": snap_ts,
            "ok": ok,
            "score": score,
            "own": own_count,
            "opp": opp_count,
            "units": strength_units,
            "strike": snap_strike,
            "reason": reason,
        })

    # Count confirmations
    # A snapshot counts as full (weight=1) if ok=True.
    # A snapshot counts as partial (weight=0.7) if delta+price pass but volume
    # is neutral (ok=False but own_count+opp_count is close to threshold).
    confirmations = []
    for snap in all_snaps:
        if snap["ok"]:
            confirmations.append(snap)
        elif STRICT_ALLOW_VOLUME_NEUTRAL:
            # partial: own>=threshold and opp>=threshold but volume was neutral
            if snap["own"] >= STRICT_MIN_SAME_STRIKES and snap["opp"] >= STRICT_MIN_OPP_STRIKES:
                confirmations.append(snap)

    if len(confirmations) < STRICT_MIN_CONFIRM_BARS:
        # for the first time after a quiet period.
        if STRICT_SPIKE_OVERRIDE_SCORE > 0 and len(confirmations) >= 1:
            last_snap = all_snaps[-1]  # current snapshot (ts itself)
            if last_snap["ok"] and last_snap["score"] >= STRICT_SPIKE_OVERRIDE_SCORE:
                reason = (
                    last_snap["reason"] +
                    f" SPIKE_OVERRIDE score={last_snap['score']}>={STRICT_SPIKE_OVERRIDE_SCORE} "
                    f"CONF={len(confirmations)}/{len(recent)}"
                )
                return True, last_snap["score"], reason
        return False, 0, (
            f"STRICT_WAIT {len(confirmations)}/{STRICT_MIN_CONFIRM_BARS} "
            f"in window {len(recent)}"
        )

    first = confirmations[0]
    last  = confirmations[-1]

    # Growing strength check: last confirmation must be >= first
    if last["units"] < first["units"]:
        return False, last["score"], (
            f"STRICT_NOT_STRONGER last_units={last['units']}<first_units={first['units']}"
        )

    reason = (
        last["reason"] +
        f" STRIKE={last['strike']} CONF={len(confirmations)}/{len(recent)} "
        f"FIRST_UNITS={first['units']} LAST_UNITS={last['units']} "
        f"FIRST_SCORE={first['score']} LAST_SCORE={last['score']}"
    )
    return True, last["score"], reason

def opposite_side(side):
    return "pe" if side == "ce" else "ce"
# 
# CORE FUNCTIONS
# 
def _finite_pct(value):
    if value is None or pd.isna(value) or not np.isfinite(float(value)):
        return 0.0
    return float(value)

def _count_true(mask):
    return int(mask.fillna(False).sum())

def _snapshot_pct_change(curr, prev):
    if curr is None or prev is None or pd.isna(curr) or pd.isna(prev):
        return 0.0
    prev = float(prev)
    if prev == 0:
        return 0.0
    return ((float(curr) - prev) / abs(prev)) * 100

def build_snapshot_analysis(full):
    """
    Per timestamp OI read for display after each complete snapshot.
    """
    analysis = {}
    prev_state = {"ce": None, "pe": None}
    state_streak = {"ce": 0, "pe": 0}
    prev_strong_side = None
    strong_streak = 0
    timestamp_list = list(pd.Index(full["timestamp"].unique()).sort_values())
    snap_by_ts = {ts: snap.copy() for ts, snap in full.groupby("timestamp", sort=True)}

    def side_snapshot(snap, side):
        p = snap[f"{side}_p_pct"].map(_finite_pct)
        v = snap[f"{side}_v_pct"].map(_finite_pct)
        d = snap[f"{side}_d_pct"].map(_finite_pct)
        g = snap[f"{side}_g_pct"].map(_finite_pct)
        n = max(1, len(snap))

        long_build = _count_true((p >= OI_BUILD_MIN_BUY_P_PCT) & (v >= OI_BUILD_MIN_BUY_V_PCT) & (d > 0))
        short_cover = _count_true((p >= OI_BUILD_MIN_BUY_P_PCT) & (v < 0))
        unwind = _count_true((p <= -OI_BUILD_MIN_SELL_P_DROP_PCT) & (v < 0))
        sell_build = _count_true((p <= -OI_BUILD_MIN_SELL_P_DROP_PCT) & (v >= OI_BUILD_MIN_BUY_V_PCT))
        delta_up = _count_true(d > 0)
        gamma_up = _count_true(g > 0)
        volume_drop = _count_true(v < 0)
        volume_surge = _count_true(v >= 30)
        price_up = _count_true(p > 0)
        price_down = _count_true(p < 0)
        gamma_hot = _count_true(g >= 20)

        avg_p = float(p.mean())
        avg_v = float(v.mean())
        avg_d = float(d.mean())
        avg_g = float(g.mean())
        max_v_idx = v.idxmax() if len(v) else None
        max_v = float(v.loc[max_v_idx]) if max_v_idx is not None and np.isfinite(float(v.loc[max_v_idx])) else 0.0
        max_v_strike = float(snap.loc[max_v_idx, "strike"]) if max_v_idx is not None else 0.0

        if long_build >= 2:
            state = "LONG_BUILD"
            label = f"fresh long build {long_build}/{n} (price+volume+delta up)"
            strength = long_build * 3 + delta_up + gamma_up
        elif short_cover >= 3:
            state = "SHORT_COVER"
            label = f"short covering {short_cover}/{n} (price up, volume down; not fresh build)"
            strength = short_cover * 2 + delta_up
        elif unwind >= 3:
            state = "UNWINDING"
            label = f"unwinding/buyers exiting {unwind}/{n} (price+volume down)"
            strength = -unwind * 3
        elif sell_build >= 3:
            state = "SELL_BUILD"
            label = f"sell pressure {sell_build}/{n} (price down, volume up)"
            strength = -sell_build * 2
        elif delta_up >= 3 and avg_p > 0:
            state = "SLOW_ACCUM"
            label = f"slow accumulation {delta_up}/{n} (delta+price improving)"
            strength = delta_up + gamma_up
        else:
            state = "MIXED"
            label = f"mixed/no clear build (Pavg {avg_p:.2f}%, Vavg {avg_v:.2f}%)"
            strength = 0

        return {
            "state": state,
            "label": label,
            "strength": strength,
            "long_build": long_build,
            "short_cover": short_cover,
            "unwind": unwind,
            "sell_build": sell_build,
            "delta_up": delta_up,
            "gamma_up": gamma_up,
            "avg_p": avg_p,
            "avg_v": avg_v,
            "avg_d": avg_d,
            "avg_g": avg_g,
            "n": n,
            "volume_drop": volume_drop,
            "volume_surge": volume_surge,
            "price_up": price_up,
            "price_down": price_down,
            "gamma_hot": gamma_hot,
            "max_v": max_v,
            "max_v_strike": max_v_strike,
        }

    def observation_points(ce, pe):
        points = []
        for side_name, info in (("PE", pe), ("CE", ce)):
            if info["volume_drop"] >= 4 and info["price_down"] >= 3:
                points.append(f"{side_name} weakening: volume dropping {info['volume_drop']}/{info['n']}, price negative {info['price_down']}/{info['n']}")
            if info["long_build"] >= 2:
                points.append(f"{side_name} accumulation: fresh build {info['long_build']}/{info['n']}, volume surge near {info['max_v_strike']:.0f}")
            elif info["short_cover"] >= 3:
                points.append(f"{side_name} short covering: price up with volume drop {info['short_cover']}/{info['n']}")
            if info["gamma_hot"] >= 3:
                points.append(f"{side_name} gamma confirms momentum on {info['gamma_hot']} strikes")
        if pe["long_build"] >= 2 and ce["long_build"] == 0 and ce["volume_surge"] < 3:
            points.append("CE buying pressure weak; PE rebuild has cleaner confirmation")
        if ce["long_build"] >= 2 and pe["long_build"] == 0 and pe["volume_surge"] < 3:
            points.append("PE buying pressure weak; CE rebuild has cleaner confirmation")
        return points[:4]

    def side_is_fading(info):
        return (
            info["state"] in ("UNWINDING", "SELL_BUILD")
            or (info["volume_drop"] >= 4 and info["price_down"] >= 3)
            or info["avg_p"] < -1
        )

    def dominance_from_snap(snap):
        ce = side_snapshot(snap, "ce")
        pe = side_snapshot(snap, "pe")
        ce_has_build = ce["state"] in ("LONG_BUILD", "SLOW_ACCUM")
        pe_has_build = pe["state"] in ("LONG_BUILD", "SLOW_ACCUM")
        ce_has_unwind = ce["state"] in ("UNWINDING", "SELL_BUILD") or ce["unwind"] >= 3 or ce["avg_p"] < -3
        pe_has_unwind = pe["state"] in ("UNWINDING", "SELL_BUILD") or pe["unwind"] >= 3 or pe["avg_p"] < -3

        if ce["strength"] > pe["strength"] + 3 and ce_has_build and pe_has_unwind:
            return "CE", "CE dominant"
        if pe["strength"] > ce["strength"] + 3 and pe_has_build and ce_has_unwind:
            return "PE", "PE dominant"
        if ce_has_unwind and (pe_has_build or pe["avg_p"] > 0 or pe["delta_up"] >= 3):
            return "PE", "PE dominant"
        if pe_has_unwind and (ce_has_build or ce["avg_p"] > 0 or ce["delta_up"] >= 3):
            return "CE", "CE dominant"
        return "MIXED", "mixed/no dominance"

    def timeframe_snap(curr_snap, prev_snap):
        prev_by_strike = {float(r["strike"]): r for _, r in prev_snap.iterrows()}
        out = curr_snap.copy()
        for idx, row in out.iterrows():
            prev_row = prev_by_strike.get(float(row["strike"]))
            for side, prefix in (("ce", "call"), ("pe", "put")):
                if prev_row is None:
                    out.at[idx, f"{side}_d_pct"] = 0.0
                    out.at[idx, f"{side}_g_pct"] = 0.0
                    out.at[idx, f"{side}_v_pct"] = 0.0
                    out.at[idx, f"{side}_p_pct"] = 0.0
                    continue
                out.at[idx, f"{side}_d_pct"] = _snapshot_pct_change(row[f"{prefix}_delta"], prev_row[f"{prefix}_delta"])
                out.at[idx, f"{side}_g_pct"] = _snapshot_pct_change(row[f"{prefix}_gamma"], prev_row[f"{prefix}_gamma"])
                out.at[idx, f"{side}_v_pct"] = _snapshot_pct_change(row[f"{prefix}_volume"], prev_row[f"{prefix}_volume"])
                out.at[idx, f"{side}_p_pct"] = _snapshot_pct_change(row[f"{prefix}_price"], prev_row[f"{prefix}_price"])
        return out

    def timeframe_verdicts(ts, curr_snap, curr_ce, curr_pe):
        verdicts = []
        for label, minutes in (("5m", 5), ("15m", 15), ("1h", 60)):
            anchor_limit = ts - pd.Timedelta(minutes=minutes)
            anchors = [old_ts for old_ts in timestamp_list if old_ts <= anchor_limit]
            if not anchors:
                verdicts.append((label, "MIXED", "not enough data"))
                continue
            anchor_snap = snap_by_ts[anchors[-1]]
            side, verdict = dominance_from_snap(timeframe_snap(curr_snap, anchor_snap))
            if side == "CE" and side_is_fading(curr_ce):
                side, verdict = "MIXED", "CE fading now"
            elif side == "PE" and side_is_fading(curr_pe):
                side, verdict = "MIXED", "PE fading now"
            verdicts.append((label, side, verdict))
        return verdicts

    for ts, snap in full.groupby("timestamp", sort=True):
        ce = side_snapshot(snap, "ce")
        pe = side_snapshot(snap, "pe")

        for side, info in (("ce", ce), ("pe", pe)):
            if info["state"] == prev_state[side] and info["state"] != "MIXED":
                state_streak[side] += 1
            else:
                state_streak[side] = 1
                prev_state[side] = info["state"]

        ce_has_build = ce["state"] in ("LONG_BUILD", "SLOW_ACCUM")
        pe_has_build = pe["state"] in ("LONG_BUILD", "SLOW_ACCUM")
        ce_has_unwind = ce["state"] in ("UNWINDING", "SELL_BUILD") or ce["unwind"] >= 3 or ce["avg_p"] < -3
        pe_has_unwind = pe["state"] in ("UNWINDING", "SELL_BUILD") or pe["unwind"] >= 3 or pe["avg_p"] < -3

        if pe["strength"] > ce["strength"] + 3 and pe_has_build and ce_has_unwind:
            strong_side = "PE"
        elif ce["strength"] > pe["strength"] + 3 and ce_has_build and pe_has_unwind:
            strong_side = "CE"
        else:
            strong_side = "MIXED"

        if strong_side == prev_strong_side and strong_side != "MIXED":
            strong_streak += 1
        else:
            strong_streak = 1
            prev_strong_side = strong_side

        def side_text(side_name, info):
            side = side_name.lower()
            notes = [info["label"]]
            if state_streak[side] >= 3 and info["state"] != "MIXED":
                notes.append(f"consistency: same {info['state'].lower()} for {state_streak[side]} snapshots")
            elif info["state"] != "MIXED":
                notes.append("consistency: early, needs more snapshots")
            if info["long_build"] and info["short_cover"]:
                notes.append(f"fresh build {info['long_build']} vs covering {info['short_cover']}")
            if info["avg_p"] > 0 and info["avg_v"] < 0:
                notes.append("price up but volume down = short covering, not fresh long")
            if info["avg_p"] < 0 and info["avg_v"] < 0:
                notes.append("price+volume down = unwinding/exit")
            if side == "ce" and info["state"] in ("LONG_BUILD", "SLOW_ACCUM") and not pe_has_unwind:
                notes.append("opposite PE unwind missing")
            if side == "pe" and info["state"] in ("LONG_BUILD", "SLOW_ACCUM") and not ce_has_unwind:
                notes.append("opposite CE unwind missing")
            return "; ".join(notes[:4])

        final_side = strong_side

        if strong_side == "PE":
            final = "PE stronger than CE from current vs past snapshot context; bearish pressure"
        elif strong_side == "CE":
            final = "CE stronger than PE from current vs past snapshot context; bullish pressure"
        elif ce_has_unwind and (pe_has_build or pe["avg_p"] > 0 or pe["delta_up"] >= 3):
            final_side = "PE"
            final = "CE unwinding while PE is improving; bearish pressure, PE side favored"
        elif pe_has_unwind and (ce_has_build or ce["avg_p"] > 0 or ce["delta_up"] >= 3):
            final_side = "CE"
            final = "PE unwinding while CE is improving; bullish pressure, CE side favored"
        elif ce_has_build and not pe_has_unwind:
            final = "CE build seen, but PE not unwinding; do not mark CE stronger yet"
        elif pe_has_build and not ce_has_unwind:
            final = "PE build seen, but CE not unwinding; do not mark PE stronger yet"
        else:
            final = "mixed/no clear side; wait for handoff"

        if strong_side != "MIXED":
            if strong_streak >= 3:
                final += f"; verdict consistency {strong_streak} snapshots"
            else:
                final += "; verdict early"
        if ce["state"] == "UNWINDING":
            final += "; CE unwinding"
        if pe["state"] == "UNWINDING":
            final += "; PE unwinding"

        pe_text = side_text("PE", pe)
        ce_text = side_text("CE", ce)
        analysis[ts] = {
            "pe": pe_text,
            "ce": ce_text,
            "final": final,
            "final_side": final_side,
            "pe_points": analysis_points(pe_text),
            "ce_points": analysis_points(ce_text),
            "final_points": analysis_points(final),
            "timeframes": timeframe_verdicts(ts, snap, ce, pe),
            "observations": observation_points(ce, pe),
        }

    return analysis

def print_snapshot_analysis(ts, snapshot_analysis):
    info = snapshot_analysis.get(ts)
    if not info:
        return
    print(f"  {bold('SNAPSHOT ANALYSIS')} {dim(str(ts)[:19])}")
    print(f"    {red('PE side:')} {info['pe']}")
    print(f"    {green('CE side:')} {info['ce']}")
    print(f"    {yellow('Final:')} {info['final']}")

def analysis_points(text, max_points=2):
    points = [p.strip() for p in str(text).split(";") if p.strip()]
    return points[:max_points] or ["mixed/no clear side"]

def _compact_note(text):
    text = str(text).strip()
    compact_map = (
        ("mixed/no clear build", "mixed/no clear"),
        ("fresh long build", "fresh build"),
        ("short covering", "short covering"),
        ("unwinding/buyers exiting", "unwinding/exit"),
        ("price up but volume down = short covering, not fresh long", "price up, vol down"),
        ("price+volume down = unwinding/exit", "price+vol down"),
        ("consistency: same", "consistent"),
        ("consistency: early, needs more snapshots", "early, need confirm"),
        ("consistency: needs more snapshots", "need confirm"),
        ("opposite PE unwind missing", "PE unwind missing"),
        ("opposite CE unwind missing", "CE unwind missing"),
        ("PE stronger than CE from current vs past snapshot context", "PE stronger vs CE"),
        ("CE stronger than PE from current vs past snapshot context", "CE stronger vs PE"),
        ("bearish pressure", "bearish pressure"),
        ("bullish pressure", "bullish pressure"),
        ("wait for handoff", "wait handoff"),
    )
    for old, new in compact_map:
        if text.startswith(old):
            text = text.replace(old, new, 1)
            break
    if text.startswith("mixed/no clear") and "(" in text:
        text = text.split("(", 1)[0].strip()
    return text

def _wrap_words(text, width):
    text = str(text).strip()
    if not text:
        return [""]
    words = text.split()
    rows = []
    current = ""
    for word in words:
        if len(word) > width:
            if current:
                rows.append(current)
                current = ""
            while len(word) > width:
                rows.append(word[:width])
                word = word[width:]
            if word:
                current = word
            continue
        candidate = word if not current else current + " " + word
        if len(candidate) <= width:
            current = candidate
        else:
            rows.append(current)
            current = word
    if current:
        rows.append(current)
    return rows

def _point_rows(number, text, width):
    text = _compact_note(text)
    prefix = f"{number}. "
    if "(" in text and text.endswith(")"):
        main, detail = text.split("(", 1)
        rows = []
        for idx, part in enumerate(_wrap_words(main.strip(), width - len(prefix))):
            rows.append((prefix if idx == 0 else "   ") + part)
        for part in _wrap_words("(" + detail.strip(), width - 3):
            rows.append("   " + part)
        return rows
    rows = []
    for idx, part in enumerate(_wrap_words(text, width - len(prefix))):
        rows.append((prefix if idx == 0 else "   ") + part)
    return rows

def snapshot_analysis_column(ts, snapshot_analysis, line_no, width=24):
    info = snapshot_analysis.get(ts)
    if not info:
        return ""

    pe_points = info.get("pe_points") or analysis_points(info.get("pe", ""))
    ce_points = info.get("ce_points") or analysis_points(info.get("ce", ""))
    final_points = info.get("final_points") or analysis_points(info.get("final", ""))

    final_side = info.get("final_side")
    final_color = red if final_side == "PE" else green if final_side == "CE" else yellow

    panel = []
    panel.append((red, "PE side:"))
    for idx, point in enumerate(pe_points[:2], 1):
        panel.extend((red, row) for row in _point_rows(idx, point, width))

    panel.append((green, "CE side:"))
    for idx, point in enumerate(ce_points[:2], 1):
        panel.extend((green, row) for row in _point_rows(idx, point, width))

    panel.append((final_color, "Final:"))
    for idx, point in enumerate(final_points[:2], 1):
        panel.extend((final_color, row) for row in _point_rows(idx, point, width))

    for label, side, verdict in info.get("timeframes", []):
        row_color = red if side == "PE" else green if side == "CE" else yellow
        panel.append((row_color, f"{label}: {verdict}"))

    for point in info.get("observations", []):
        row_color = red if point.startswith("PE") or "CE buying pressure weak" in point else green if point.startswith("CE") or "PE buying pressure weak" in point else yellow
        for row in _wrap_words(point, width):
            panel.append((row_color, row))

    if final_side in ("PE", "CE"):
        direction = "downside bias" if final_side == "PE" else "upside bias"
        panel.append((final_color, "Key: " + direction))

    if 1 <= line_no <= len(panel):
        color, text = panel[line_no - 1]
        return color(text)
    return ""


def pct_change_col(series):
    return (series.pct_change(fill_method=None) * 100).round(2)

def rolling_pct_rank(series, window=20):
    return series.rolling(window, min_periods=1).apply(
        lambda x: (x[-1] > x[:-1]).mean() * 100 if len(x) > 1 else 50.0,
        raw=True
    )

def compute_strength_score(delta_rank, gamma_rank, volume_rank, price_rank):
    return (
        0.30 * delta_rank +
        0.30 * gamma_rank +
        0.25 * volume_rank +
        0.15 * price_rank
    ).round(1)

def cross_candle_trend(series, n):
    def slope_sign(x):
        if len(x) < 2: return 0
        xs = np.arange(len(x))
        slope = np.polyfit(xs, x, 1)[0]
        if slope > 0:  return 1
        if slope < 0:  return -1
        return 0
    return series.rolling(n, min_periods=2).apply(slope_sign, raw=True)

def cross_candle_pct(series, n):
    return ((series - series.shift(n)) / series.shift(n).abs() * 100).round(2)

def safe_float(x):
    try:
        v = float(x)
        return None if np.isnan(v) else v
    except Exception:
        return None

# 
# XLSX EXPORT   mimics console colours as cell fills / fonts
# 
def _export_xlsx(full, layer1_rows, layer2_rows, opp_rows, sides, col_map, save_cols, timestamps):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XLSide
    from openpyxl.utils import get_column_letter

    #  Palette 
    HDR_BG      = PatternFill("solid", start_color="1F4E79")
    HDR_FONT    = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(name="Consolas", size=9)
    BOLD_FONT   = Font(name="Consolas", size=9, bold=True)
    ALT_FILL    = PatternFill("solid", start_color="1A1A2E")   # dark alt row
    DARK_FILL   = PatternFill("solid", start_color="16213E")   # base dark
    BULL_FILL   = PatternFill("solid", start_color="1A3A1A")   # dark green
    BEAR_FILL   = PatternFill("solid", start_color="3A1A1A")   # dark red
    SPIKE_FILL  = PatternFill("solid", start_color="2A2A0A")   # dark yellow

    # text colours
    C_BRIGHT_GREEN = "92FF92"
    C_GREEN        = "00CC00"
    C_BRIGHT_RED   = "FF9292"
    C_RED          = "FF4444"
    C_YELLOW       = "FFFF00"
    C_CYAN         = "00FFFF"
    C_DIM          = "888888"
    C_WHITE        = "FFFFFF"

    thin = XLSide(style="thin", color="333355")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(color=C_WHITE): return Font(name="Consolas", bold=True,  color=color, size=9)
    def bfont(color=C_WHITE, bold=False): return Font(name="Consolas", bold=bold, color=color, size=9)

    def pct_color(val, extreme=20.0):
        """Return (text_str, hex_color) for a pct value."""
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return "n/a", C_DIM
        sign = "^" if val > 0 else "v" if val < 0 else " "
        txt  = f"{sign}{abs(val):.2f}%"
        if val > extreme:    clr = C_BRIGHT_GREEN
        elif val > 0:        clr = C_GREEN
        elif val < -extreme: clr = C_BRIGHT_RED
        elif val < 0:        clr = C_RED
        else:                clr = C_DIM
        return txt, clr

    def score_color(score):
        if score >= 75: return C_BRIGHT_GREEN
        if score >= 50: return C_GREEN
        if score >= 25: return C_YELLOW
        return C_RED

    def score_label(score):
        if score >= 75: return "STRONG"
        if score >= 50: return "MODERATE"
        if score >= 25: return "WEAK"
        return "VERY WK"

    def write_cell(ws, row, col, value, fg=C_WHITE, bg=None, bold=False, align="center"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(name="Consolas", color=fg, bold=bold, size=9)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = bdr
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        return cell

    def style_header_row(ws, cols, row=1):
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.font      = HDR_FONT
            cell.fill      = HDR_BG
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
        ws.row_dimensions[row].height = 18

    def autofit(ws, extra=3, max_w=35):
        for col_cells in ws.columns:
            best = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(best + extra, max_w)

    wb = Workbook()

    # 
    # SHEET 1  LAYER 1  Same-Candle Spikes
    # 
    ws1 = wb.active
    ws1.title       = "L1 Same-Candle Spikes"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "FFEB9C"

    l1_hdr = ["TIMESTAMP","STRIKE","SPOT","SIDE","SCORE","STRENGTH","LTP",
              "DELTA","DELTA_%","GAMMA","GAMMA_%",
              "VOLUME","VOLUME_%","PRICE","PRICE_%","IV"]
    style_header_row(ws1, l1_hdr)

    for ri, (ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv) in enumerate(layer1_rows, 2):
        row_bg = "1A3A1A" if side == "ce" else "3A1A1A"
        sc     = score_color(score)
        sl     = score_label(score)
        dp_t, dp_c = pct_color(safe_float(dp))
        gp_t, gp_c = pct_color(safe_float(gp))
        vp_t, vp_c = pct_color(safe_float(vp), 50)
        pp_t, pp_c = pct_color(safe_float(pp))

        write_cell(ws1, ri, 1,  str(ts)[:19],             C_DIM,          row_bg)
        write_cell(ws1, ri, 2,  round(stk,1),              C_CYAN,         row_bg)
        write_cell(ws1, ri, 3,  round(spot,2),             C_CYAN,         row_bg)
        write_cell(ws1, ri, 4,  side.upper(),              C_BRIGHT_GREEN if side=="ce" else C_BRIGHT_RED, row_bg, bold=True)
        write_cell(ws1, ri, 5,  round(score,1),            sc,             row_bg, bold=True)
        write_cell(ws1, ri, 6,  sl,                        sc,             row_bg)
        write_cell(ws1, ri, 7,  round(float(pv),2) if safe_float(pv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 8,  round(float(dv),4) if safe_float(dv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 9,  dp_t,                      dp_c,           row_bg)
        write_cell(ws1, ri, 10, round(float(gv),6) if safe_float(gv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 11, gp_t,                      gp_c,           row_bg)
        write_cell(ws1, ri, 12, int(vv) if safe_float(vv) is not None else None, C_DIM, row_bg)
        write_cell(ws1, ri, 13, vp_t,                      vp_c,           row_bg)
        write_cell(ws1, ri, 14, round(float(pv),2) if safe_float(pv) is not None else None, C_WHITE, row_bg)
        write_cell(ws1, ri, 15, pp_t,                      pp_c,           row_bg)
        write_cell(ws1, ri, 16, round(float(iv),2) if safe_float(iv) is not None else None, C_DIM, row_bg)

    ws1.freeze_panes = "A2"
    autofit(ws1)

    # 
    # SHEET 2  LAYER 2  Cross-Candle Trend
    # 
    ws2 = wb.create_sheet("L2 Cross-Candle Trend")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00BFFF"

    l2_hdr = ["TIMESTAMP","STRIKE","SPOT","SIDE","DIRECTION","SCORE",
              f"DELTA_%_{CROSS_CANDLES}c", f"GAMMA_%_{CROSS_CANDLES}c",
              f"VOLUME_%_{CROSS_CANDLES}c", f"PRICE_%_{CROSS_CANDLES}c"]
    style_header_row(ws2, l2_hdr)

    for ri, (ts, stk, spot, side, direction, score, d_xp, g_xp, v_xp, p_xp) in enumerate(layer2_rows, 2):
        is_bull = direction == "BULLISH"
        row_bg  = "1A3A1A" if is_bull else "3A1A1A"
        sc      = score_color(score)
        dir_c   = C_BRIGHT_GREEN if is_bull else C_BRIGHT_RED
        dxt, dxc = pct_color(safe_float(d_xp), 30)
        gxt, gxc = pct_color(safe_float(g_xp), 30)
        vxt, vxc = pct_color(safe_float(v_xp), 50)
        pxt, pxc = pct_color(safe_float(p_xp), 20)

        write_cell(ws2, ri, 1,  str(ts)[:19],   C_DIM,  row_bg)
        write_cell(ws2, ri, 2,  round(stk,1),   C_CYAN, row_bg)
        write_cell(ws2, ri, 3,  round(spot,2),  C_CYAN, row_bg)
        write_cell(ws2, ri, 4,  side.upper(),   C_BRIGHT_GREEN if side=="CE" else C_BRIGHT_RED, row_bg, bold=True)
        write_cell(ws2, ri, 5,  direction,      dir_c,  row_bg, bold=True)
        write_cell(ws2, ri, 6,  round(score,1), sc,     row_bg, bold=True)
        write_cell(ws2, ri, 7,  dxt,            dxc,    row_bg)
        write_cell(ws2, ri, 8,  gxt,            gxc,    row_bg)
        write_cell(ws2, ri, 9,  vxt,            vxc,    row_bg)
        write_cell(ws2, ri, 10, pxt,            pxc,    row_bg)

    ws2.freeze_panes = "A2"
    autofit(ws2)

    # 
    # SHEET 3  LAYER 3  Cross-Strike Confirmation
    # 
    ws3 = wb.create_sheet("L3 Cross-Strike Confirm")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "9900FF"

    offsets   = [i for i in range(-STRIKES_NEARBY, STRIKES_NEARBY+1) if i != 0]
    l3_hdr    = ["TIMESTAMP","ATM","SPOT","SIDE","ATM SCORE"] + \
                [f"ATM{'+' if i>0 else ''}{i}" for i in offsets]
    style_header_row(ws3, l3_hdr)

    all_strikes_local = sorted(full["strike"].unique())

    ri = 2
    for ts in timestamps:
        ts_data = full[full["timestamp"] == ts]
        spot    = ts_data["spot"].iloc[0]
        atm_idx = (ts_data["strike"] - spot).abs().idxmin()
        atm_stk = ts_data.loc[atm_idx, "strike"]
        try:
            atm_pos = all_strikes_local.index(atm_stk)
        except ValueError:
            continue

        for side in sides:
            atm_score_s = ts_data.loc[ts_data["strike"]==atm_stk, f"{side}_score"]
            if atm_score_s.empty: continue
            atm_score = atm_score_s.iloc[0]
            if atm_score < STRENGTH_PCT: continue

            row_bg = "1A1A3A"
            write_cell(ws3, ri, 1, str(ts)[:19],  C_DIM,  row_bg)
            write_cell(ws3, ri, 2, round(atm_stk,1), C_CYAN, row_bg)
            write_cell(ws3, ri, 3, round(spot,2),    C_CYAN, row_bg)
            write_cell(ws3, ri, 4, side.upper(), C_BRIGHT_GREEN if side=="ce" else C_BRIGHT_RED, row_bg, bold=True)
            write_cell(ws3, ri, 5, round(atm_score,1), score_color(atm_score), row_bg, bold=True)

            for ci, offset in enumerate(offsets, 6):
                ni = atm_pos + offset
                if 0 <= ni < len(all_strikes_local):
                    nstk = all_strikes_local[ni]
                    row  = ts_data[ts_data["strike"]==nstk]
                    sc   = row[f"{side}_score"].iloc[0] if not row.empty else None
                else:
                    sc = None
                if sc is not None:
                    write_cell(ws3, ri, ci, round(sc,1), score_color(sc), row_bg)
                else:
                    write_cell(ws3, ri, ci, "n/a", C_DIM, row_bg)
            ri += 1

    ws3.freeze_panes = "A2"
    autofit(ws3)

    # 
    # SHEET 4  LAYER 4  Opposite-Side Weak Confirmation
    # 
    ws4 = wb.create_sheet("L4 Opp-Side Confirm")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "FF4500"

    l4_hdr = ["TIMESTAMP","STRIKE","SPOT","SIGNAL",
              "CE SCORE","PE SCORE",
              "CE D%","CE G%","PE D%","PE G%","CONFIDENCE"]
    style_header_row(ws4, l4_hdr)

    for ri, (ts, stk, spot, signal, ce_sc, pe_sc, ce_dp, ce_gp, pe_dp, pe_gp, conf) in enumerate(opp_rows, 2):
        sig_clean = strip_ansi(signal)
        is_bull   = "BULL" in sig_clean
        row_bg    = "1A3A1A" if is_bull else "3A1A1A"
        sig_c     = C_BRIGHT_GREEN if is_bull else C_BRIGHT_RED
        sig_txt   = " BULLISH CONFIRM" if is_bull else " BEARISH CONFIRM"

        ce_dp_t, ce_dp_c = pct_color(safe_float(ce_dp))
        ce_gp_t, ce_gp_c = pct_color(safe_float(ce_gp))
        pe_dp_t, pe_dp_c = pct_color(safe_float(pe_dp))
        pe_gp_t, pe_gp_c = pct_color(safe_float(pe_gp))

        write_cell(ws4, ri, 1,  str(ts)[:19],       C_DIM,             row_bg)
        write_cell(ws4, ri, 2,  round(stk,1),        C_CYAN,            row_bg)
        write_cell(ws4, ri, 3,  round(spot,2),       C_CYAN,            row_bg)
        write_cell(ws4, ri, 4,  sig_txt,             sig_c,             row_bg, bold=True)
        write_cell(ws4, ri, 5,  round(ce_sc,1),      score_color(ce_sc),row_bg, bold=True)
        write_cell(ws4, ri, 6,  round(pe_sc,1),      score_color(pe_sc),row_bg, bold=True)
        write_cell(ws4, ri, 7,  ce_dp_t,             ce_dp_c,           row_bg)
        write_cell(ws4, ri, 8,  ce_gp_t,             ce_gp_c,           row_bg)
        write_cell(ws4, ri, 9,  pe_dp_t,             pe_dp_c,           row_bg)
        write_cell(ws4, ri, 10, pe_gp_t,             pe_gp_c,           row_bg)
        write_cell(ws4, ri, 11, round(conf,1),        score_color(min(conf,100)), row_bg, bold=True)

    ws4.freeze_panes = "A2"
    autofit(ws4)

    # 
    # SHEET 5  Full Data
    # 
    ws_fd = wb.create_sheet("Full Data")
    ws_fd.sheet_view.showGridLines = False
    ws_fd.sheet_properties.tabColor = "444444"

    existing = [c for c in save_cols if c in full.columns]
    style_header_row(ws_fd, existing)

    for ri2, row in enumerate(full[existing].itertuples(index=False), 2):
        bg = "1A1A2E" if ri2 % 2 == 0 else "16213E"
        for ci2, val in enumerate(row, 1):
            if hasattr(val, "isoformat"):
                v = val.isoformat()
            elif isinstance(val, (bool, np.bool_)):
                v = str(val)
            elif isinstance(val, float) and np.isnan(val):
                v = None
            else:
                v = val
            write_cell(ws_fd, ri2, ci2, v, C_WHITE, bg)

    ws_fd.freeze_panes = "A2"
    autofit(ws_fd)

    # 
    # Set dark background for all sheets' tab area via sheet background
    # 
    wb.save(OUT_XLSX)


# 
# MAIN
# 
def run_single_analysis(csv_path=None, analysis_date=None):
    _STRICT_BEST_CACHE.clear()
    csv_path = csv_path or CSV_PATH
    analysis_date = analysis_date or ANALYSIS_DATE

    csv = Path(csv_path)
    if not csv.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}\nRun 1_db_to_csv.py first.")

    print(bold(cyan(f"\n  Loading {csv_path} ")))
    df = pd.read_csv(csv, parse_dates=["timestamp"])
    print("CSV date range:", df["timestamp"].min(), "to", df["timestamp"].max())


    # Keep only selected date from the CSV
    # Keep only selected date from the CSV
    df = df[df["timestamp"].dt.date == pd.to_datetime(analysis_date).date()].copy()
    df.reset_index(drop=True, inplace=True)

    # Stop script clearly if selected date is not available in CSV
    if df.empty:
        print(red(f"No rows found for ANALYSIS_DATE = {analysis_date}"))
        print(yellow("Use one date from Available dates printed above."))
        return {
            "csv": Path(csv_path).name,
            "date": str(analysis_date),
            "trades": 0,
            "profit_count": 0,
            "loss_count": 0,
            "flat_count": 0,
            "total_pnl": 0.0,
            "best_move_pnl": 0.0,
            "bad_move_pnl": 0.0,
            "trade_details": [],
        }

    df.sort_values(["timestamp", "strike"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    #  Session time filter 
    before = len(df)

    df = df[df["timestamp"].dt.time.between(SESSION_START, SESSION_END)].copy()
    df.reset_index(drop=True, inplace=True)

    # Keep only ATM 100 points for each timestamp
    df["atm_strike"] = (df["spot"] / STRIKE_STEP).round() * STRIKE_STEP

    df = df[
        (df["strike"] >= df["atm_strike"] - ATM_RANGE_POINTS) &
        (df["strike"] <= df["atm_strike"] + ATM_RANGE_POINTS)
    ].copy()

    df.drop(columns=["atm_strike"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    after  = len(df)
    print(f"  {dim('Time filter:')} {SESSION_START.strftime('%H:%M')}  {SESSION_END.strftime('%H:%M')}  "
          f"{dim('Rows:')} {before:,}  {bright_green(str(after))}")

    sides   = ["ce","pe"] if SIDE == "both" else [SIDE]
    col_map = {
        "ce": {"delta":"call_delta","gamma":"call_gamma",
               "volume":"call_volume","price":"call_price","iv":"call_iv"},
        "pe": {"delta":"put_delta", "gamma":"put_gamma",
               "volume":"put_volume","price":"put_price","iv":"put_iv"},
    }

    print(f"  {dim('Timestamps:')} {df['timestamp'].nunique()}  "
          f"{dim('Strikes:')} {df['strike'].nunique()}  "
          f"{dim('Sides:')} {SIDE.upper()}  "
          f"{dim('OI print interval:')} {OI_PRINT_INTERVAL}  "
          f"{dim('X-candle window:')} {CROSS_CANDLES}  "
          f"{dim('Nearby strikes:')} {STRIKES_NEARBY}\n")

    #  Per-strike feature engineering 
    all_strikes = sorted(df["strike"].unique())
    records     = []

    for strike in all_strikes:
        sdf = df[df["strike"] == strike].copy().sort_values("timestamp").reset_index(drop=True)
        if len(sdf) < 3:
            continue

        for side in sides:
            cm = col_map[side]
            d  = sdf[cm["delta"]].copy()
            g  = sdf[cm["gamma"]].copy()
            v  = sdf[cm["volume"]].copy()
            p  = sdf[cm["price"]].copy()

            sdf[f"{side}_d_pct"]   = pct_change_col(d)
            sdf[f"{side}_g_pct"]   = pct_change_col(g)
            sdf[f"{side}_v_pct"]   = pct_change_col(v)
            sdf[f"{side}_p_pct"]   = pct_change_col(p)

            sdf[f"{side}_d_xpct"]  = cross_candle_pct(d, CROSS_CANDLES)
            sdf[f"{side}_g_xpct"]  = cross_candle_pct(g, CROSS_CANDLES)
            sdf[f"{side}_v_xpct"]  = cross_candle_pct(v, CROSS_CANDLES)
            sdf[f"{side}_p_xpct"]  = cross_candle_pct(p, CROSS_CANDLES)

            sdf[f"{side}_d_trend"] = cross_candle_trend(d, CROSS_CANDLES)
            sdf[f"{side}_g_trend"] = cross_candle_trend(g, CROSS_CANDLES)
            sdf[f"{side}_v_trend"] = cross_candle_trend(v, CROSS_CANDLES)
            sdf[f"{side}_p_trend"] = cross_candle_trend(p, CROSS_CANDLES)

            sdf[f"{side}_d_rank"]  = rolling_pct_rank(d)
            sdf[f"{side}_g_rank"]  = rolling_pct_rank(g)
            sdf[f"{side}_v_rank"]  = rolling_pct_rank(v)
            sdf[f"{side}_p_rank"]  = rolling_pct_rank(p)

            sdf[f"{side}_score"]   = compute_strength_score(
                sdf[f"{side}_d_rank"], sdf[f"{side}_g_rank"],
                sdf[f"{side}_v_rank"], sdf[f"{side}_p_rank"]
            )

            sdf[f"{side}_same_spike"] = (
                (sdf[f"{side}_d_rank"] >= STRENGTH_PCT) &
                (sdf[f"{side}_g_rank"] >= STRENGTH_PCT)
            )
            # Display-only OI buildup row:
            # Big moves can start with option price + volume expansion before
            # delta/gamma ranks become strong. Keep these candles visible in
            # the OI buildup table for breakdowns like 2024-01-02 09:54-10:00.
            sdf[f"{side}_price_volume_build"] = (
                (sdf[f"{side}_p_pct"] >= OI_TABLE_PRICE_PCT) &
                (sdf[f"{side}_v_pct"] >= OI_TABLE_VOLUME_PCT)
            )
            # Also keep the opposite-side weakness visible. During a PE move,
            # CE rows often have price falling and should be shown as weakness
            # context instead of disappearing from the table.
            sdf[f"{side}_price_drop_weakness"] = (
                sdf[f"{side}_p_pct"] <= OI_TABLE_PRICE_DROP_PCT
            )
            sdf[f"{side}_cross_bull"] = (
                (sdf[f"{side}_d_trend"] == 1) &
                (sdf[f"{side}_g_trend"] == 1) &
                (sdf[f"{side}_v_trend"] == 1)
            )
            sdf[f"{side}_cross_bear"] = (
                (sdf[f"{side}_d_trend"] == -1) &
                (sdf[f"{side}_g_trend"] == -1) &
                (sdf[f"{side}_v_trend"] == -1)
            )

        records.append(sdf)

    if not records:
        print(yellow(f"No usable option rows after filters for ANALYSIS_DATE = {analysis_date}"))
        return {
            "csv": Path(csv_path).name,
            "date": str(analysis_date),
            "trades": 0,
            "profit_count": 0,
            "loss_count": 0,
            "flat_count": 0,
            "total_pnl": 0.0,
            "best_move_pnl": 0.0,
            "bad_move_pnl": 0.0,
            "trade_details": [],
        }

    full       = pd.concat(records, ignore_index=True).sort_values(["timestamp","strike"])
    snapshot_analysis = build_snapshot_analysis(full)
    timestamps = sorted(full["timestamp"].unique())

    W  = 177   # console width - wide enough for all columns plus compact snapshot analysis
    W3 = W + STRIKES_NEARBY * 14

    # 
    # LAYER 1  SAME-CANDLE SPIKE REPORT
    # 
    header(f"SNAPSHOT OI BUILDUP  every {OI_PRINT_INTERVAL}, CE first then PE", W)
    print(f"  {dim('Layer 1 threshold: runtime percentile ')} {bold(str(STRENGTH_PCT))}th\n")

    print(
        f"  {dim('Also showing price-volume buildup rows: P% >=')} "
        f"{bold(str(OI_TABLE_PRICE_PCT))}   {dim('V% >=')} "
        f"{bold(str(OI_TABLE_VOLUME_PCT))}   "
        f"{dim('and weakness rows: P% <=')} {bold(str(OI_TABLE_PRICE_DROP_PCT))}\n"
    )

    h1 = (f"  {'TIMESTAMP':<19}  {'STRIKE':>8}  {'SPOT':>8}  {'SIDE':>4}  "
          f"{'SCORE':>5}  {'STRENGTH':>8}  {'LTP':>8}  "
          f"{'DELTA':>10} {'D CHG%':>9}  {'GAMMA':>10} {'G CHG%':>9}  "
          f"{'VOLUME':>10} {'V%':>9}  {'PRICE':>8} {'P%':>9}  {'IV':>7}  {'ANALYSIS':<24}")
    print(bold(h1))
    sep(W)
    
    
    snapshot_rows = []
    layer1_rows = []
    display_full = full[full["timestamp"].map(is_print_interval_timestamp)].copy()
    for side in sides:
        for _, r in display_full.iterrows():
            snapshot_rows.append((r["timestamp"], r["strike"], r["spot"], side,
                                  r[f"{side}_score"],
                                  r[col_map[side]["delta"]], r[f"{side}_d_pct"],
                                  r[col_map[side]["gamma"]], r[f"{side}_g_pct"],
                                  r[col_map[side]["volume"]],r[f"{side}_v_pct"],
                                  r[col_map[side]["price"]], r[f"{side}_p_pct"],
                                  r[col_map[side]["iv"]]))

        sub = full[
            full[f"{side}_same_spike"] |
            full[f"{side}_price_volume_build"] |
            full[f"{side}_price_drop_weakness"]
        ].copy()
        for _, r in sub.iterrows():
            layer1_rows.append((r["timestamp"], r["strike"], r["spot"], side,
                                r[f"{side}_score"],
                                r[col_map[side]["delta"]], r[f"{side}_d_pct"],
                                r[col_map[side]["gamma"]], r[f"{side}_g_pct"],
                                r[col_map[side]["volume"]],r[f"{side}_v_pct"],
                                r[col_map[side]["price"]], r[f"{side}_p_pct"],
                                r[col_map[side]["iv"]]))

    # Same timestamp  all CE first  all PE next
    snapshot_rows.sort(key=lambda x: (x[0], 0 if x[3] == "ce" else 1, x[1]))
    layer1_rows.sort(key=lambda x: (x[0], 0 if x[3] == "ce" else 1, x[1]))
    entry_source_rows = layer1_rows
    entry_source_rows.sort(key=lambda x: (x[0], 0 if x[3] == "ce" else 1, x[1]))

    prev_snapshot_ts = None
    snapshot_line_no = 0
    for ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv in snapshot_rows:
        is_first_snapshot_row = ts != prev_snapshot_ts
        if prev_snapshot_ts is not None and ts != prev_snapshot_ts:
            print()
            snapshot_line_no = 0
        prev_snapshot_ts = ts
        snapshot_line_no += 1
        analysis_txt = snapshot_analysis_column(ts, snapshot_analysis, snapshot_line_no)

        sc     = f"{score:5.1f}"
        lbl    = strength_label(score)
        side_c = bright_green("CE") if side == "ce" else bright_red("PE")
        dv_d   = dv
        gv_d   = gv
        print(
            f"  {dim(str(ts)[:19])}  "
            f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
            f"{bold(sc)}  {lbl}  {dim(f'{pv:>8.2f}')}  "
            f"{rjust(dim(f'{dv_d:>10.4f}'), 10)} {rjust(color_pct(dp), 9)}  "
            f"{rjust(dim(f'{gv_d:>10.6f}'), 10)} {rjust(color_pct(gp), 9)}  "
            f"{rjust(dim(f'{vv:>10,.0f}'), 10)} {rjust(color_pct(vp, 50), 9)}  "
            f"{dim(f'{pv:>8.2f}')} {rjust(color_pct(pp), 9)}  "
            f"{dim(f'{iv:>7.2f}%')}  {analysis_txt:<24}"
        )

    if not snapshot_rows:
        print(f"  {dim('No snapshot rows found.')}")
    sep(W)

    # TOP ENTRIES TABLE  best entry candles from Layer 1 data
    # Logic: delta up + gamma up + volume up + price up
    # Extra rule: one best row per timestamp, so duplicate same-time entries are removed
    # 
    header("TOP ENTRIES  Delta + Gamma + Volume + Price increasing", W)
    pe_total = 0
    pe_strict_reject = 0
    pe_pass = 0
    top_rows = []

    for row in entry_source_rows:
        ts, stk, spot, side, score, dv, dp, gv, gp, vv, vp, pv, pp, iv = row

        if side == "pe":
            pe_total += 1

        r = full[
            (full["timestamp"] == ts) &
            (full["strike"] == stk)
        ].iloc[0]

        if ts.time() >= ENTRY_CUTOFF:
            continue

        strict_entry_ok, strict_entry_score, strict_entry_reason = strict_cross_entry_ok(
            full, ts, stk, side
        )
        entry_allowed = strict_entry_ok

        if entry_allowed:

            if side == "pe":
                pe_pass += 1
            
            entry_score = 100 + strict_entry_score + abs(dp) + min(max(vp, 0), 500) * 0.10 + max(pp, 0)

            entry_price = pv  # option buy price at entry candle

            exit_ts, exit_price, _, exit_reason = find_exit_after_entry(
                full,
                ts,
                stk,
                side,
                entry_price,
                col_map,
                strict_entry=True,
            )

            entry_price = pv
            pnl_points = exit_price - entry_price if exit_price is not None else np.nan

            best_move, best_time, best_price = compute_strike_best_move(
                full, ts, stk, side, entry_price, col_map
            )
            bad_move, bad_time, bad_price = compute_strike_bad_move(
                full, ts, stk, side, entry_price, col_map
            )
            top_rows.append((
            ts, stk, spot, side,
            entry_price, exit_ts, exit_price, pnl_points,
            best_move, best_time, best_price,
            bad_move, bad_time, bad_price,
            entry_score, score, dp, gp, vp, pp,
            exit_reason,
            f"STRICT_ENTRY {strict_entry_reason} STOP={STOP_LOSS_PTS}"
            ))
        
        else:
            if side == "pe":
                pe_strict_reject += 1

    # First sort by ENTRY SCORE so duplicate same-time signals keep the strongest row.
    top_rows.sort(key=lambda x: x[14], reverse=True)

    # Remove duplicate timestamps: keep only strongest row from same timestamp.
    unique_rows = []
    seen_ts = set()

    for r in top_rows:
        ts = str(r[0])[:19]
        if ts in seen_ts:
            continue
        seen_ts.add(ts)
        unique_rows.append(r)

    # Print entries in proper time order, not score order.
    unique_rows.sort(key=lambda x: x[0])
    overlapping_skip_count = 0
    side_switch_count = 0
    if not ALLOW_OVERLAPPING_TRADES:
        single_trade_rows = []
        active_until = None
        active_side = None

        for r in unique_rows:
            entry_ts  = r[0]
            exit_ts   = r[5]
            new_side  = r[3]

            if active_until is not None and entry_ts <= active_until:
                # Same side still active → skip as before
                if new_side == active_side:
                    overlapping_skip_count += 1
                    continue

                # OPPOSITE side fires while trade is active → side switch:
                # Close the active trade at this new entry candle price, then
                # open the new opposite-side trade immediately.
                if ALLOW_SIDE_SWITCH:
                    # Find the price of the active trade's strike at entry_ts
                    prev = single_trade_rows[-1]
                    prev_strike = prev[1]
                    prev_side   = prev[3]
                    cm_prev     = col_map[prev_side]
                    switch_rows = full[
                        (full["timestamp"] == entry_ts) &
                        (full["strike"] == prev_strike)
                    ]
                    if not switch_rows.empty:
                        switch_exit_price = float(switch_rows.iloc[0][cm_prev["price"]])
                        switch_pnl = switch_exit_price - float(prev[4])
                        # Rebuild best/bad move up to switch point
                        bm, bt, bp = compute_strike_best_move(full, prev[0], prev_strike, prev_side, prev[4], col_map)
                        bdm, bdt, bdp = compute_strike_bad_move(full, prev[0], prev_strike, prev_side, prev[4], col_map)
                        # Replace the last row with updated exit
                        updated_prev = (
                            prev[0], prev[1], prev[2], prev[3],
                            prev[4],                        # entry_price
                            entry_ts, switch_exit_price, switch_pnl,
                            bm, bt, bp,
                            bdm, bdt, bdp,
                            prev[14], prev[15],
                            prev[16], prev[17], prev[18], prev[19],
                            f"Exit: side-switch to {new_side.upper()} at {str(entry_ts)[:19]}",
                            prev[21],
                        )
                        single_trade_rows[-1] = updated_prev

                    # Now open the new opposite-side trade
                    side_switch_count += 1
                    single_trade_rows.append(r)
                    active_until = exit_ts if exit_ts is not None and not pd.isna(exit_ts) else entry_ts
                    active_side  = new_side
                    continue
                else:
                    overlapping_skip_count += 1
                    continue

            single_trade_rows.append(r)
            active_until = exit_ts if exit_ts is not None and not pd.isna(exit_ts) else entry_ts
            active_side  = new_side

        unique_rows = single_trade_rows

    print(f"  {dim('Entry model:')} {bold('STRICT CROSS-STRIKE ONLY')}   "
          f"{dim('Stop loss:')} {bold(str(STOP_LOSS_PTS) + ' pts')}   "
          f"{dim('Force exit:')} {bold(FORCE_EXIT.strftime('%H:%M'))}")
    print(f"  {dim('Entry criteria:')} "
          f"{dim('own side D/P/V strength + opposite side D/P weakness; repeated confirmation must strengthen')}")
    print(f"  {dim('Strict params:')} "
          f"{dim('V% >=')} {bold(str(STRICT_MIN_VOLUME_PCT))}   "
          f"{dim('D% >=')} {bold(str(STRICT_MIN_DELTA_PCT))}   "
          f"{dim('P% >=')} {bold(str(STRICT_MIN_PRICE_PCT))}   "
          f"{dim('Same/Opp >=')} {bold(str(STRICT_MIN_SAME_STRIKES) + '/' + str(STRICT_MIN_OPP_STRIKES))}   "
          f"{dim('Confirm:')} {bold(str(STRICT_MIN_CONFIRM_BARS) + '/' + str(STRICT_CONFIRM_WINDOW))}   "
          f"{dim('Exit bars:')} {bold(str(STRICT_EXIT_CONFIRM_BARS))}   "
          f"{dim('Surge ratio:')} {bold(str(STRICT_EXIT_SURGE_RATIO))}x   "
          f"{dim('Spike override:')} {bold(str(STRICT_SPIKE_OVERRIDE_SCORE) if STRICT_SPIKE_OVERRIDE_SCORE > 0 else 'OFF')}")
    if ALLOW_OVERLAPPING_TRADES:
        print(f"  {dim('Single active trade rule:')} {bold('OFF')}   "
              f"{dim('Overlapping entries allowed')}")
    else:
        switch_txt = (bright_green(f"Side switches: {side_switch_count}") if side_switch_count > 0
                      else dim(f"Side switches: {side_switch_count}"))
        print(f"  {dim('Single active trade rule:')} {bold('ON')}   "
              f"{dim('Overlapping entries skipped:')} {bold(str(overlapping_skip_count))}   "
              f"{switch_txt}   "
              f"{dim('Side-switch:')} {bold('ON' if ALLOW_SIDE_SWITCH else 'OFF')}")
    print()

    h_top = (
        f"  {'ENTRY TIME':<19}  {'EXIT TIME':<19}  {'STRIKE':>8}  {'SPOT':>8}  {'SIDE':>4}  "
        f"{'ENTRY':>8}  {'EXIT':>8}  {'PNL':>8}  {'BEST MOVE':>10}  {'BAD MOVE':>9}  {'BEST TIME':<19}  "
        f"{'ENTRY SCR':>9}  {'SYS SCR':>7}  "
        f"{'D%':>9}  {'G%':>9}  {'V%':>9}  {'P%':>9}  {'EXIT REASON':<28}"
    )
    print(bold(h_top))
    sep(W)

    for ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move, best_time, best_price, bad_move, bad_time, bad_price, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason in unique_rows[:TOP_N]:
        side_c = bright_green("CE") if side == "ce" else bright_red("PE")

        pnl_txt = bright_green(f"{pnl_points:>8.2f}") if pnl_points >= 0 else bright_red(f"{pnl_points:>8.2f}")
        best_txt = bright_green(f"{best_move:>10.2f}") if best_move >= 0 else bright_red(f"{best_move:>10.2f}")
        bad_txt = bright_green(f"{bad_move:>9.2f}") if bad_move >= 0 else bright_red(f"{bad_move:>9.2f}")

        print(
            f"  {dim(str(ts)[:19])}  "
            f"{dim(str(exit_ts)[:19])}  "
            f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
            f"{entry_price:>8.2f}  {exit_price:>8.2f}  {pnl_txt}  {best_txt}  {bad_txt}  "
            f"{dim(str(best_time)[:19])}  "
            f"{bold(f'{entry_score:>9.1f}')}  {bold(f'{sys_score:>7.1f}')}  "
            f"{rjust(color_pct(dp), 9)}  "
            f"{rjust(color_pct(gp), 9)}  "
            f"{rjust(color_pct(vp, 50), 9)}  "
            f"{rjust(color_pct(pp), 9)}  "
            f"{exit_reason:<28}  {reason}"
        )
    pe_rows = [x for x in unique_rows if x[3] == "pe"]

    if pe_rows:
        print()
        header("PE ENTRIES ONLY  same format", W)
        print(bold(h_top))
        sep(W)

        for ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move, best_time, best_price, bad_move, bad_time, bad_price, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason in pe_rows[:TOP_N]:
            side_c = bright_red("PE")

            pnl_txt = bright_green(f"{pnl_points:>8.2f}") if pnl_points >= 0 else bright_red(f"{pnl_points:>8.2f}")
            best_txt = bright_green(f"{best_move:>10.2f}") if best_move >= 0 else bright_red(f"{best_move:>10.2f}")
            bad_txt = bright_green(f"{bad_move:>9.2f}") if bad_move >= 0 else bright_red(f"{bad_move:>9.2f}")

            print(
                f"  {dim(str(ts)[:19])}  "
                f"{dim(str(exit_ts)[:19])}  "
                f"{cyan(f'{stk:>8.1f}')}  {cyan(f'{spot:>8.2f}')}  {side_c}  "
                f"{entry_price:>8.2f}  {exit_price:>8.2f}  {pnl_txt}  {best_txt}  {bad_txt}  "
                f"{dim(str(best_time)[:19])}  "
                f"{bold(f'{entry_score:>9.1f}')}  {bold(f'{sys_score:>7.1f}')}  "
                f"{rjust(color_pct(dp), 9)}  "
                f"{rjust(color_pct(gp), 9)}  "
                f"{rjust(color_pct(vp, 50), 9)}  "
                f"{rjust(color_pct(pp), 9)}  "
                f"{exit_reason:<28}  {reason}"
            )

        sep(W)
    if not unique_rows:
        print(f"  {dim('No top entries found.')}")

    sep(W)
    print("\n========== STRICT ENTRY DEBUG ==========")
    print("PE candidate rows      :", pe_total)
    print("PE strict rejected     :", pe_strict_reject)
    print("PE strict accepted     :", pe_pass)
    print("Overlapping skipped    :", overlapping_skip_count)
    print("========================================")


    
    # 
    # SUMMARY
    # 
    header("SUMMARY", W)

    option_pnls = [
        float(r[7]) for r in unique_rows
        if r[7] is not None and not pd.isna(r[7])
    ]
    total_option_pnl = round(sum(option_pnls), 2) if option_pnls else 0.0
    profit_count = sum(1 for pnl in option_pnls if pnl > 0)
    loss_count = sum(1 for pnl in option_pnls if pnl < 0)
    flat_count = sum(1 for pnl in option_pnls if pnl == 0)
    avg_option_pnl = round(total_option_pnl / len(option_pnls), 2) if option_pnls else 0.0

    # Summary after removing Layer 2, Layer 3, Layer 4 tables
    print(f"  {dim('Session window              :')} {bold(SESSION_START.strftime('%H:%M'))}  {bold(SESSION_END.strftime('%H:%M'))}")
    print(f"  {dim('Layer 1  Same-candle spikes:')} {bold(str(len(layer1_rows)))}")
    print(f"  {dim('Top entries shown           :')} {bold(str(min(len(unique_rows), TOP_N)))}")
    print(f"  {dim('Option trades total         :')} {bold(str(len(option_pnls)))}")
    print(f"  {dim('Option profit count         :')} {bold(str(profit_count))}")
    print(f"  {dim('Option loss count           :')} {bold(str(loss_count))}")
    print(f"  {dim('Option flat count           :')} {bold(str(flat_count))}")
    total_pnl_txt = bright_green(f"{total_option_pnl:.2f}") if total_option_pnl >= 0 else bright_red(f"{total_option_pnl:.2f}")
    avg_pnl_txt = bright_green(f"{avg_option_pnl:.2f}") if avg_option_pnl >= 0 else bright_red(f"{avg_option_pnl:.2f}")
    print(f"  {dim('Option total PNL            :')} {bold(total_pnl_txt)}")
    print(f"  {dim('Option average PNL          :')} {bold(avg_pnl_txt)}")
    print()

    #  Save CSV 
    save_cols = ["timestamp","strike","spot"]
    for side in sides:
        cm = col_map[side]
        save_cols += [
            cm["delta"], f"{side}_d_pct", f"{side}_d_xpct", f"{side}_d_rank", f"{side}_d_trend",
            cm["gamma"], f"{side}_g_pct", f"{side}_g_xpct", f"{side}_g_rank", f"{side}_g_trend",
            cm["volume"],f"{side}_v_pct", f"{side}_v_xpct", f"{side}_v_rank", f"{side}_v_trend",
            cm["price"], f"{side}_p_pct", f"{side}_p_xpct", f"{side}_p_rank", f"{side}_p_trend",
            cm["iv"],    f"{side}_score", f"{side}_same_spike",
            f"{side}_cross_bull", f"{side}_cross_bear",
        ]

    trade_details = []
    for r in unique_rows:
        ts, stk, spot, side, entry_price, exit_ts, exit_price, pnl_points, best_move, best_time, best_price, bad_move, bad_time, bad_price, entry_score, sys_score, dp, gp, vp, pp, exit_reason, reason = r
        trade_details.append({
            "entry_time": str(ts)[:19],
            "exit_time": str(exit_ts)[:19],
            "side": side.upper(),
            "strike": float(stk),
            "entry": float(entry_price),
            "exit": float(exit_price),
            "pnl": float(pnl_points),
            "best_move": float(best_move),
            "best_time": str(best_time)[:19],
            "bad_move": float(bad_move),
            "bad_time": str(bad_time)[:19],
            "entry_score": float(entry_score),
            "sys_score": float(sys_score),
            "d_pct": float(dp),
            "g_pct": float(gp),
            "v_pct": float(vp),
            "p_pct": float(pp),
            "exit_reason": str(exit_reason),
        })

    return {
        "csv": Path(csv_path).name,
        "date": str(analysis_date),
        "trades": len(option_pnls),
        "profit_count": profit_count,
        "loss_count": loss_count,
        "flat_count": flat_count,
        "total_pnl": total_option_pnl,
        "best_move_pnl": round(sum(float(r[8]) for r in unique_rows if r[8] is not None and not pd.isna(r[8])), 2),
        "bad_move_pnl": round(sum(float(r[11]) for r in unique_rows if r[11] is not None and not pd.isna(r[11])), 2),
        "trade_details": trade_details,
    }

def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Options flow strength analyser. No args keeps hardcoded CSV/date "
            "full-output mode. --monthly enables compact date-by-date summary mode."
        )
    )
    parser.add_argument("--csv", "--csv-path", dest="csv_path",
                        help="Run one CSV file instead of hardcoded CSV_PATH.")
    parser.add_argument("--date", "--analysis-date", dest="analysis_date",
                        help="Run one date instead of hardcoded ANALYSIS_DATE.")
    parser.add_argument("--monthly", action="store_true",
                        help="Compact mode: run weekly expiry CSVs for a selected month/year.")
    parser.add_argument("--month",
                        help="Month name or number for --monthly, e.g. jan, january, 1.")
    parser.add_argument("--year", type=int,
                        help="Year for --monthly, e.g. 2024.")
    parser.add_argument("--oi-data", "--oi-data-mode", choices=["back", "live"], default=OI_DATA_MODE,
                        help="OI CSV source: back=niftyoptiondata, live=Strategy3MAGreeks\\live_Saidata_weeklyexp.")
    parser.add_argument("--csv-dir",
                        help="Folder containing weekly oi_YYYY_MM_DD.csv files. Overrides --oi-data for monthly mode.")
    parser.add_argument("--print-interval", "--oi-print-interval", type=parse_print_interval_arg, default=OI_PRINT_INTERVAL,
                        help="Snapshot OI buildup print interval: 5min default, or 1min to print every candle.")
    parser.add_argument("--no-next-week", action="store_true",
                        help="Monthly mode: do not include first 7 days of next month.")
    parser.add_argument("--strict-volume-pct", type=float,
                        help="Minimum own-side volume percentage. Default: 80.")
    parser.add_argument("--strict-delta-pct", type=float,
                        help="Minimum own-side delta percentage and opposite-side delta drop. Default: 2.")
    parser.add_argument("--strict-price-pct", type=float,
                        help="Minimum own-side price percentage and opposite-side price drop. Default: 3.")
    parser.add_argument("--strict-same-strikes", type=int,
                        help="Minimum confirming same-side nearby strikes. Default: 3.")
    parser.add_argument("--strict-opp-strikes", type=int,
                        help="Minimum confirming opposite-side nearby strikes. Default: 2.")
    parser.add_argument("--strict-window", type=int,
                        help="Recent snapshot confirmation window. Default: 5.")
    parser.add_argument("--strict-confirm-bars", type=int,
                        help="Minimum confirmations inside the window. Default: 3.")
    parser.add_argument("--strict-exit-bars", type=int,
                        help="Opposite-side confirmations needed to exit. Default: 2.")
    parser.add_argument("--exit-surge-ratio", type=float,
                        help="Early exit if 2nd opposite confirmation score > this ratio × 1st score. "
                             "Default: 1.5. Use 0 to disable early surge exit. Try 1.2, 1.5, 2.0 to tune.")
    parser.add_argument("--spike-override-score", type=float,
                        help="Enter immediately (skipping full confirm-bars requirement) if current snapshot "
                             "composite score >= this AND at least 1 prior confirm exists in the window. "
                             "Default: 60. Use 0 to disable spike override entirely. "
                             "Lower = catches more moves early. Higher = only fires on extreme spikes. "
                             "Try 40, 60, 80, 100 to tune.")
    parser.add_argument("--no-side-switch", action="store_true",
                        help="Disable side-switch: opposite-side signals while a trade is active are skipped "
                             "instead of closing the current trade and flipping. Default: side-switch ON.")
    parser.add_argument("--strict-require-volume", action="store_true",
                        help="Require volume threshold on own-side confirmations; otherwise delta+price with non-negative volume can count.")
    parser.add_argument("--console-xlsx", nargs="?", const=OUT_CONSOLE_XLSX,
                        help="Dump the exact colored console output to an Excel file. "
                             f"Default file: {OUT_CONSOLE_XLSX}.")
    return parser.parse_args()

def apply_arg_config(args):
    global OI_DATA_MODE
    global OI_PRINT_INTERVAL
    global STRICT_MIN_VOLUME_PCT, STRICT_MIN_DELTA_PCT
    global STRICT_MIN_PRICE_PCT, STRICT_MIN_SAME_STRIKES, STRICT_MIN_OPP_STRIKES
    global STRICT_CONFIRM_WINDOW, STRICT_MIN_CONFIRM_BARS, STRICT_ALLOW_VOLUME_NEUTRAL
    global STRICT_EXIT_CONFIRM_BARS, STRICT_EXIT_SURGE_RATIO, STRICT_SPIKE_OVERRIDE_SCORE
    global ALLOW_SIDE_SWITCH

    OI_DATA_MODE = args.oi_data
    OI_PRINT_INTERVAL = args.print_interval

    if args.strict_volume_pct is not None:
        STRICT_MIN_VOLUME_PCT = args.strict_volume_pct
    if args.strict_delta_pct is not None:
        STRICT_MIN_DELTA_PCT = args.strict_delta_pct
    if args.strict_price_pct is not None:
        STRICT_MIN_PRICE_PCT = args.strict_price_pct
    if args.strict_same_strikes is not None:
        STRICT_MIN_SAME_STRIKES = args.strict_same_strikes
    if args.strict_opp_strikes is not None:
        STRICT_MIN_OPP_STRIKES = args.strict_opp_strikes
    if args.strict_window is not None:
        STRICT_CONFIRM_WINDOW = args.strict_window
    if args.strict_confirm_bars is not None:
        STRICT_MIN_CONFIRM_BARS = args.strict_confirm_bars
    if args.strict_exit_bars is not None:
        STRICT_EXIT_CONFIRM_BARS = args.strict_exit_bars
    if args.exit_surge_ratio is not None:
        STRICT_EXIT_SURGE_RATIO = args.exit_surge_ratio
    if args.spike_override_score is not None:
        STRICT_SPIKE_OVERRIDE_SCORE = args.spike_override_score
    if args.no_side_switch:
        ALLOW_SIDE_SWITCH = False
    if args.strict_require_volume:
        STRICT_ALLOW_VOLUME_NEUTRAL = False

def print_compact_monthly_summary(results, month, year, start_date, end_date):
    header(f"MONTHLY SUMMARY - {month_name(month)} {year}", 112)
    print(f"  Date range: {start_date} to {end_date}")
    print(f"  {'CSV FILE':<20} {'DATE':<10} {'TRADES':>6} {'PROFIT':>6} {'LOSS':>6} {'FLAT':>5} {'PNL':>10} {'BEST MOVE':>10} {'BAD MOVE':>10}")
    sep(112)

    total_trades = 0
    total_profit = 0
    total_loss = 0
    total_flat = 0
    total_pnl = 0.0
    total_best = 0.0
    total_bad = 0.0

    for r in results:
        pnl_txt = bright_green(f"{r['total_pnl']:>10.2f}") if r["total_pnl"] >= 0 else bright_red(f"{r['total_pnl']:>10.2f}")
        best_txt = bright_green(f"{r['best_move_pnl']:>10.2f}") if r["best_move_pnl"] >= 0 else bright_red(f"{r['best_move_pnl']:>10.2f}")
        bad_txt = bright_green(f"{r['bad_move_pnl']:>10.2f}") if r["bad_move_pnl"] >= 0 else bright_red(f"{r['bad_move_pnl']:>10.2f}")
        print(f"  {r['csv']:<20} {r['date']:<10} {r['trades']:>6} {r['profit_count']:>6} "
              f"{r['loss_count']:>6} {r['flat_count']:>5} {pnl_txt} {best_txt} {bad_txt}")

        if r.get("trade_details"):
            print(f"    {'ENTRY TIME':<19} {'EXIT TIME':<19} {'SIDE':<4} {'STRIKE':>8} "
                  f"{'ENTRY':>8} {'EXIT':>8} {'PNL':>9} {'BEST':>9} {'BAD':>9} {'BEST TIME':<19} "
                  f"{'D%':>8} {'G%':>8} {'V%':>8} {'P%':>8}  EXIT REASON")
            for t in r["trade_details"]:
                trade_pnl = bright_green(f"{t['pnl']:>9.2f}") if t["pnl"] >= 0 else bright_red(f"{t['pnl']:>9.2f}")
                trade_best = bright_green(f"{t['best_move']:>9.2f}") if t["best_move"] >= 0 else bright_red(f"{t['best_move']:>9.2f}")
                trade_bad = bright_green(f"{t['bad_move']:>9.2f}") if t["bad_move"] >= 0 else bright_red(f"{t['bad_move']:>9.2f}")
                print(f"    {t['entry_time']:<19} {t['exit_time']:<19} {t['side']:<4} "
                      f"{t['strike']:>8.1f} {t['entry']:>8.2f} {t['exit']:>8.2f} "
                      f"{trade_pnl} {trade_best} {trade_bad} {t['best_time']:<19} "
                      f"{t['d_pct']:>7.2f}% {t['g_pct']:>7.2f}% {t['v_pct']:>7.2f}% "
                      f"{t['p_pct']:>7.2f}%  {t['exit_reason']}")
            print()

        total_trades += r["trades"]
        total_profit += r["profit_count"]
        total_loss += r["loss_count"]
        total_flat += r["flat_count"]
        total_pnl += r["total_pnl"]
        total_best += r["best_move_pnl"]
        total_bad += r["bad_move_pnl"]

    sep(112)
    total_pnl_txt = bright_green(f"{total_pnl:>10.2f}") if total_pnl >= 0 else bright_red(f"{total_pnl:>10.2f}")
    total_best_txt = bright_green(f"{total_best:>10.2f}") if total_best >= 0 else bright_red(f"{total_best:>10.2f}")
    total_bad_txt = bright_green(f"{total_bad:>10.2f}") if total_bad >= 0 else bright_red(f"{total_bad:>10.2f}")
    print(f"  {'TOTAL':<20} {'':<10} {total_trades:>6} {total_profit:>6} {total_loss:>6} "
          f"{total_flat:>5} {total_pnl_txt} {total_best_txt} {total_bad_txt}")
    print()

def run_monthly(args):
    if not args.month or not args.year:
        raise SystemExit("--monthly needs both --month and --year")

    month = parse_month(args.month)
    start_date, end_date = monthly_date_range(
        args.year,
        month,
        include_next_week=not args.no_next_week,
    )

    csv_dir = Path(args.csv_dir) if args.csv_dir else (LIVE_OI_DATA_DIR if args.oi_data == "live" else BACK_OI_DATA_DIR)
    date_to_csv = {}
    for csv_file in sorted(csv_dir.glob("oi_*.csv")):
        file_date = date_from_oi_filename(csv_file)
        if not file_date:
            continue

        # A month-end trading date can live inside the next weekly expiry CSV.
        # Scan one extra week of expiry files, but only keep trading dates that
        # are inside the requested monthly date range.
        if file_date < start_date or file_date > (pd.Timestamp(end_date) + pd.Timedelta(days=7)).date():
            continue

        for trade_date in trading_dates_in_csv(csv_file):
            if not (start_date <= trade_date <= end_date):
                continue
            current = date_to_csv.get(trade_date)
            if current is None:
                date_to_csv[trade_date] = (csv_file, file_date)
                continue
            _, current_file_date = current
            if file_date >= trade_date and (current_file_date < trade_date or file_date < current_file_date):
                date_to_csv[trade_date] = (csv_file, file_date)

    csv_files = [(csv_file, trade_date) for trade_date, (csv_file, _) in sorted(date_to_csv.items())]

    if not csv_files:
        print(red(f"No trading dates found in {csv_dir} for {start_date} to {end_date}"))
        return []

    results = []
    for csv_file, trade_date in csv_files:
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            result = run_single_analysis(str(csv_file), str(trade_date))
        results.append(result)

    print_compact_monthly_summary(results, month, args.year, start_date, end_date)
    return results

def main():
    if len(sys.argv) == 1:
        run_single_analysis()
        return

    args = parse_args()
    apply_arg_config(args)
    console_capture = None
    old_stdout = sys.stdout
    if args.console_xlsx:
        console_capture = TeeConsole(sys.stdout)
        sys.stdout = console_capture

    try:
        if args.monthly:
            run_monthly(args)
            return

        csv_path = Path(args.csv_path or CSV_PATH)
        if not csv_path.is_absolute():
            csv_path = (LIVE_OI_DATA_DIR if args.oi_data == "live" else BACK_OI_DATA_DIR) / csv_path

        run_single_analysis(str(csv_path), args.analysis_date or ANALYSIS_DATE)
    finally:
        if console_capture is not None:
            sys.stdout = old_stdout
            export_console_xlsx(console_capture.getvalue(), args.console_xlsx)


if __name__ == "__main__":
    main()

